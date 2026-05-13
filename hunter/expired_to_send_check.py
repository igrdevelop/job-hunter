"""
hunter/expired_to_send_check.py — Check to_send.xlsx URLs for expired job offers.

Used by the Telegram /check_expired command.

Strategy:
  - Works on a COPY (to_send_checking.xlsx) so the original is never touched
    while the user may have it open in Excel or LibreOffice Calc.
  - Fetches URLs in parallel: global semaphore + per-domain semaphore + per-domain delay.
  - After the check, /apply_expired replaces the original with the copy.
"""

import asyncio
import logging
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Callable, Awaitable
from urllib.parse import urlparse

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill

from hunter.config import (
    PROJECT_DIR,
    EXPIRED_CHECK_CONCURRENCY,
    EXPIRED_CHECK_DOMAIN_LIMIT,
    EXPIRED_CHECK_DOMAIN_DELAY,
)
from hunter.expired_check import is_job_expired

logger = logging.getLogger(__name__)

TO_SEND_PATH  = PROJECT_DIR / "to_send.xlsx"
CHECKING_PATH = PROJECT_DIR / "to_send_checking.xlsx"

# Send a progress update every N completed rows
PROGRESS_EVERY = 5


# ── Per-domain rate limiting ──────────────────────────────────────────────────

def _domain(url: str) -> str:
    return urlparse(url).hostname or url


class _DomainLimiter:
    """Per-domain semaphore + delay to avoid hammering a single site."""

    def __init__(self, domain_limit: int, domain_delay: float) -> None:
        self._limit = domain_limit
        self._delay = domain_delay
        self._sems: dict[str, asyncio.Semaphore] = {}
        self._delays: dict[str, asyncio.Lock] = {}

    def _get(self, dom: str):
        if dom not in self._sems:
            self._sems[dom] = asyncio.Semaphore(self._limit)
            self._delays[dom] = asyncio.Lock()
        return self._sems[dom], self._delays[dom]

    async def fetch(self, url: str, global_sem: asyncio.Semaphore) -> str:
        from job_fetch import fetch_job_text
        dom = _domain(url)
        dom_sem, dom_lock = self._get(dom)

        async with global_sem:
            async with dom_sem:
                result = await asyncio.to_thread(fetch_job_text, url)
                # Enforce per-domain delay via a lock so concurrent domain
                # requests don't overlap their sleep windows
                async with dom_lock:
                    await asyncio.sleep(self._delay)
                return result


# ── Core check logic ──────────────────────────────────────────────────────────

async def run_check(
    progress_cb: Callable[[str], Awaitable[None]] | None = None,
) -> dict:
    """
    Copy to_send.xlsx → to_send_checking.xlsx, fetch each URL in parallel,
    check for expiry. Marks EXPIRED rows in the COPY only.
    Call apply_check() afterwards to replace the original.

    Calls progress_cb(message) every PROGRESS_EVERY completed rows if provided.

    Returns:
        {
            "total":   int,
            "alive":   int,
            "expired": list[dict],   # {row, company, title, url}
            "errors":  list[dict],   # {row, company, title, error}
            "marked":  int,
        }
    """
    if not TO_SEND_PATH.exists():
        raise FileNotFoundError(f"to_send.xlsx not found: {TO_SEND_PATH}")

    # Work on a fresh copy — never touch the original while an editor may have it open
    await asyncio.to_thread(shutil.copy2, TO_SEND_PATH, CHECKING_PATH)
    logger.info("[check_expired] Working copy created: %s", CHECKING_PATH)

    wb = openpyxl.load_workbook(CHECKING_PATH)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def _col(name: str) -> int:
        return headers.index(name) + 1

    url_col     = _col("URL")
    sent_col    = _col("Sent")
    company_col = _col("Company")
    title_col   = _col("Job Title")

    # Collect rows to check (skip already EXPIRED)
    to_check: list[dict] = []
    for row_num in range(2, ws.max_row + 1):
        url     = ws.cell(row_num, url_col).value
        company = ws.cell(row_num, company_col).value or ""
        title   = ws.cell(row_num, title_col).value or ""
        sent    = ws.cell(row_num, sent_col).value

        if not url:
            continue
        if str(sent or "").upper() == "EXPIRED":
            continue
        to_check.append({"row": row_num, "url": str(url), "company": company, "title": title})

    total   = len(to_check)
    done    = 0
    expired_count = 0

    global_sem = asyncio.Semaphore(EXPIRED_CHECK_CONCURRENCY)
    limiter    = _DomainLimiter(EXPIRED_CHECK_DOMAIN_LIMIT, EXPIRED_CHECK_DOMAIN_DELAY)

    async def _check_one(item: dict) -> dict:
        nonlocal done, expired_count

        # jobleads.com is always Cloudflare-blocked — skip immediately, no fetch
        if "jobleads.com" in item["url"]:
            done += 1
            logger.debug("[check_expired] jobleads skip (no fetch): %s", item["url"])
            if progress_cb and done % PROGRESS_EVERY == 0:
                await progress_cb(f"⏳ {done}/{total} проверено — ⏭ истекло: {expired_count}")
            return {**item, "status": "skipped"}

        try:
            text = await limiter.fetch(item["url"], global_sem)
            if is_job_expired(text):
                status = "expired"
            else:
                status = "alive"
        except requests.exceptions.HTTPError as e:
            if e.response is not None and e.response.status_code == 404:
                status = "expired"
                item = {**item, "reason": "404"}
            else:
                status = "error"
                item = {**item, "error": str(e)}
        except Exception as e:
            from job_fetch.jobleads import JobLeadsCloudflareError
            if isinstance(e, JobLeadsCloudflareError) or "jobleads.com" in item["url"]:
                status = "alive"   # Cloudflare-blocked — uncheckable, skip silently
                logger.debug("[check_expired] jobleads skip: %s", item["url"])
            else:
                status = "error"
                item = {**item, "error": str(e)}

        done += 1
        if status == "expired":
            expired_count += 1
        if progress_cb and done % PROGRESS_EVERY == 0:
            await progress_cb(f"⏳ {done}/{total} проверено — ⏭ истекло: {expired_count}")

        return {**item, "status": status}

    # Run all checks in parallel (limited by semaphores)
    results = await asyncio.gather(*[_check_one(item) for item in to_check])

    # Apply results to the workbook copy
    expired: list[dict] = []
    errors:  list[dict] = []
    skipped: list[dict] = []
    alive = 0

    for res in results:
        if res["status"] == "expired":
            expired.append(res)
            logger.info("[check_expired] EXPIRED: %s — %s", res["company"], res["title"])
            cell = ws.cell(res["row"], sent_col, value="EXPIRED")
            cell.fill = PatternFill("solid", fgColor="FCE4D6")
            cell.font = Font(name="Calibri", size=11, color="9C0006", bold=True)
        elif res["status"] == "skipped":
            skipped.append(res)
        elif res["status"] == "alive":
            alive += 1
        else:
            errors.append(res)
            logger.warning("[check_expired] error %s: %s", res["url"], res.get("error", ""))

    # Always save the copy
    marked = len(expired)
    await asyncio.to_thread(wb.save, CHECKING_PATH)
    logger.info(
        "[check_expired] Copy saved — %d EXPIRED, %d alive, %d errors",
        marked, alive, len(errors),
    )

    return {
        "total":   total,
        "alive":   alive,
        "expired": expired,
        "errors":  errors,
        "skipped": skipped,
        "marked":  marked,
    }


# ── Apply checked copy to original ───────────────────────────────────────────

def apply_check() -> dict:
    """
    Replace to_send.xlsx with to_send_checking.xlsx (the checked copy).
    Call only after the user has closed to_send.xlsx in Excel / LibreOffice Calc.

    Returns:
        {"ok": bool, "error": str | None}
    """
    if not CHECKING_PATH.exists():
        return {
            "ok": False,
            "error": "to_send_checking.xlsx не найден — запусти /check_expired сначала.",
        }

    try:
        CHECKING_PATH.replace(TO_SEND_PATH)
        logger.info("[check_expired] Applied: %s → %s", CHECKING_PATH.name, TO_SEND_PATH.name)
    except PermissionError:
        return {"ok": False, "error": "PermissionError"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

    # Propagate EXPIRED marks to tracker.xlsx
    try:
        from hunter import to_send
        sync_result = to_send.sync_and_rebuild()
        synced = sync_result.get("synced", 0)
        logger.info("[check_expired] sync_sent: %d mark(s) synced to tracker.xlsx", synced)
        return {"ok": True, "error": None, "synced": synced}
    except Exception as e:
        logger.warning("[check_expired] sync_sent failed: %s", e)
        return {"ok": True, "error": None, "synced": 0, "sync_error": str(e)}
