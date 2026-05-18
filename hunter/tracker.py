"""
tracker.py — read/write tracker.xlsx for the hunter.

Responsibilities:
  - get_known_urls()     → set of URLs already in tracker (for dedup)
  - add_skipped(job)     → append a row with status "SKIP"
  - add_applied(...)     → append a successful generated-docs row
  - add_manual_jobleads_pending(...) → MANUAL row when JobLeads text cannot be fetched
"""

import re
import time
import uuid
from datetime import date
from pathlib import Path
from urllib.parse import urlparse, urlunparse, parse_qs, urlencode

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from hunter.config import TRACKER_PATH
from hunter.models import Job

TRACKER_HEADERS = [
    "Date", "Company", "Job Title", "Stack",
    "ATS %", "URL", "Folder", "Sent", "Re-application", "To Learn", "ID",
]
# Columns: 1 Date, 2 Company, 3 Job Title, 4 Stack, 5 ATS %, 6 URL, 7 Folder, 8 Sent, 9 Re-app, 10 To Learn, 11 ID
URL_COL_INDEX = 6       # "URL" (was wrongly 5 — that is ATS %, broke URL dedup)
COMPANY_COL_INDEX = 2   # "Company"
TITLE_COL_INDEX = 3     # "Job Title"
ATS_COL_INDEX = 5       # "ATS %" - also used for status (FAIL, SKIP)
SENT_COL_INDEX = 8      # "Sent"
ID_COL_INDEX = 11       # "ID" — short uuid4 hex, used as sync key (Google Sheets ↔ tracker)
REACT_SKIP_SENT_MARKERS = {"—", "–", "-"}

# ATS column: JobLeads detail pages are Cloudflare-blocked — user pastes description
# into Applications/.../job_posting.txt then re-runs apply on the same URL.
MANUAL_PENDING_ATS = "MANUAL"


def normalize_url(url: str) -> str:
    """Canonical form: lowercase host, strip trailing slash, drop tracking params."""
    url = url.strip()
    if not url:
        return url
    p = urlparse(url)
    host = (p.hostname or "").lower()
    path = p.path.rstrip("/") or "/"
    # Drop common tracking query params
    drop_params = {
        "utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term",
        "utm_id", "utm_term",
        "ref", "refId", "trackingId", "trk", "fbclid", "gclid",
        "campaignid", "adgroupid",
        "originToLandingJobPostings", "origin",
        # Pracuj.pl tracking params (email alerts, suggested jobs)
        "sendid", "send_date", "sug",
    }
    qs = parse_qs(p.query, keep_blank_values=False)
    clean_qs = {k: v for k, v in qs.items() if k not in drop_params}
    query = urlencode(clean_qs, doseq=True) if clean_qs else ""
    # For LinkedIn /jobs/view/{id}/ — keep only path
    if "linkedin.com" in host and "/jobs/view/" in path:
        m = re.search(r"/jobs/view/(\d+)", path)
        if m:
            path = f"/jobs/view/{m.group(1)}"
            query = ""
    # Strip /rss suffix (solid.jobs RSS feed vs regular URL are the same job)
    if path.endswith("/rss"):
        path = path[:-4] or "/"
    # Normalize JustJoin URL format: /offers/{slug} → /job-offer/{slug}
    # JustJoin changed their URL scheme; both point to the same offer.
    if "justjoin.it" in host:
        path = re.sub(r"^/offers/", "/job-offer/", path)
    return urlunparse((p.scheme, host, path, "", query, ""))


def _strip_diacritics(s: str) -> str:
    """Transliterate Polish/accented chars to ASCII equivalents.

    Note: ł/Ł (U+0142/U+0141) are NOT NFKD-decomposable, so they must be
    replaced explicitly before the standard normalization step.
    """
    import unicodedata
    s = s.replace('\u0142', 'l').replace('\u0141', 'L')   # ł → l, Ł → L
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')


def _strip_legal_suffixes(s: str) -> str:
    """Remove Polish/international legal-form suffixes in all common representations.

    Must be called AFTER lowercasing and diacritic-stripping (so 'ł' is already 'l').

    Handles both properly formatted ("Sp. z o.o.", "S.A.") and squished/CamelCase
    variants that appear in LLM-generated folder names:
      MindboxSpZOo  → mindbox
      UpvantaSpólkaZOgraniczonąOdpowiedzialnoś → upvanta  (after diacritics → spolka...)
    """
    # "sp. z o.o." and variants (spaces/dots optional) — no \b needed, dots are specific enough
    s = re.sub(r'sp\.?\s*z\.?\s*o\.?\s*o\.?', '', s)
    # "S.A." — trailing \b doesn't work after a dot, use (?![a-z]) lookahead
    s = re.sub(r's\.a\.(?![a-z])', '', s)
    # Other international forms — use \b where safe (all-alpha, no trailing dots)
    s = re.sub(r'\b(ltd|gmbh|inc|llc)\b\.?', '', s)
    # "Spółka..." / "Spolka..." — remove from this word to end of string
    # After diacritic-strip: ł→l, ó→o, so "Spółka" → "Spolka"
    s = re.sub(r'spol?ka.*', '', s)
    # Squished forms: SpZOo, SpZoo, spzoo (sp + z + oo)
    s = re.sub(r'spzo+', '', s)
    return s


def normalize_company(company: str) -> str:
    """Normalized company name for dedup.

    Produces the same key for all of these:
      - "Mindbox Sp. z o.o."
      - "MindboxSpZOo"   (LLM folder-safe form)
      - "Upvanta Spółka z o.o."
      - "UpvantaSpółkaZOgraniczonąOdpowiedzialnoś"
    """
    s = company.lower()
    s = re.sub(r'_?\d{4}-\d{2}-\d{2}(_\d+)?$', '', s)
    s = _strip_diacritics(s)
    s = _strip_legal_suffixes(s)
    s = re.sub(r'[^a-z0-9]', '', s)
    return s


def dedup_key(company: str, title: str) -> str:
    """Normalized key for company+title dedup (cross-source, cross-URL)."""
    def _norm(s: str) -> str:
        s = s.lower()
        s = re.sub(r'_?\d{4}-\d{2}-\d{2}(_\d+)?$', '', s)
        s = _strip_diacritics(s)
        s = _strip_legal_suffixes(s)
        s = re.sub(r'[^a-z0-9]', '', s)
        return s
    return _norm(company) + "|" + _norm(title)


# ── Internal helpers ──────────────────────────────────────────────────────────

def _save_with_retry(wb: openpyxl.Workbook, retries: int = 5, delay: float = 3.0) -> None:
    """Save workbook, retrying on PermissionError (file open in Excel)."""
    for attempt in range(1, retries + 1):
        try:
            wb.save(TRACKER_PATH)
            return
        except PermissionError:
            if attempt == retries:
                raise
            print(
                f"[tracker] tracker.xlsx is locked (Excel open?). "
                f"Retry {attempt}/{retries} in {delay}s..."
            )
            time.sleep(delay)


def _new_row_id() -> str:
    """Generate a short unique ID for a tracker row (8-char hex)."""
    return uuid.uuid4().hex[:8]


def _ensure_ids(ws: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Assign IDs to any existing rows that lack one. Returns True if any were added."""
    changed = False
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=ID_COL_INDEX)
        if not cell.value:
            cell.value = _new_row_id()
            changed = True
    return changed


def _load_or_create() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
        # Ensure header row has ID column (migration for existing files)
        if ws.max_column < ID_COL_INDEX:
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="2B579A")
            cell = ws.cell(row=1, column=ID_COL_INDEX, value="ID")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ID_COL_INDEX)].width = 12
        if _ensure_ids(ws):
            _save_with_retry(wb)
        return wb, ws

    # Create fresh tracker with header row
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    header_fill = PatternFill("solid", fgColor="2B579A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    widths = [12, 20, 30, 12, 8, 50, 40, 8, 16, 35, 12]

    for col, (header, width) in enumerate(zip(TRACKER_HEADERS, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    _save_with_retry(wb)
    return wb, ws


# ── Public API ────────────────────────────────────────────────────────────────

def get_known_urls() -> set[str]:
    """Return all normalized URLs stored in tracker — used for deduplication."""
    from hunter import db as _db
    return _db.get_known_norm_urls()


def get_known_company_titles() -> set[str]:
    """Return dedup_key(company, title) for all rows in tracker."""
    from hunter import db as _db
    return _db.get_known_ct_keys()


def get_sent_companies() -> set[str]:
    """Return normalized company names for rows that have Sent info."""
    from hunter import db as _db
    return {normalize_company(c) for c in _db.get_sent_company_names() if c}


def company_matches_sent(comp_key: str, sent_companies: set[str]) -> bool:
    """Fuzzy company match: 'scalo' matches 'scalowroclaw' and vice versa.

    True if comp_key is a substring of any sent company, or any sent company
    is a substring of comp_key. Minimum 3 chars to avoid false positives.
    """
    if not comp_key or len(comp_key) < 3:
        return False
    if comp_key in sent_companies:
        return True
    for sc in sent_companies:
        if len(sc) < 3:
            continue
        if comp_key in sc or sc in comp_key:
            return True
    return False


def get_failed_jobs() -> list[Job]:
    """Return Job objects for all rows with ATS status 'FAIL'."""
    from hunter import db as _db
    return [
        Job(title=r["title"], company=r["company"], location="", salary=None,
            url=r["url"], source="retry")
        for r in _db.get_by_ats("FAIL")
        if r["url"]
    ]


def remove_failed(url: str) -> None:
    """Remove a FAIL row from tracker (so it can be re-added as a proper entry)."""
    from hunter import db as _db
    norm = normalize_url(url)
    _db.delete_where(norm, "FAIL")

    if not TRACKER_PATH.exists():
        return
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    rows_to_delete = []
    for row_num in range(2, ws.max_row + 1):
        ats = str(ws.cell(row=row_num, column=ATS_COL_INDEX).value or "").strip()
        row_url = str(ws.cell(row=row_num, column=URL_COL_INDEX).value or "").strip()
        if ats == "FAIL" and normalize_url(row_url) == norm:
            rows_to_delete.append(row_num)
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)
    if rows_to_delete:
        _save_with_retry(wb)
    else:
        wb.close()


def has_successful_entry(url: str) -> bool:
    """True if tracker has a non-FAIL, non-SKIP entry for this URL (= docs were generated)."""
    return get_url_status_flags(url)["has_success"]


def get_url_status_flags(url: str) -> dict[str, bool]:
    """Return status flags for URL: successful entry and React-only skip marker."""
    from hunter import db as _db
    norm = normalize_url(url)
    rows = _db.get_by_norm_url(norm)
    has_success = False
    is_react_skip = False
    for r in rows:
        ats = r["ats"].strip().upper()
        sent = r["sent"].strip()
        if ats not in ("FAIL", "SKIP", "?", "", MANUAL_PENDING_ATS):
            has_success = True
        elif ats == "SKIP" and sent in REACT_SKIP_SENT_MARKERS:
            is_react_skip = True
        if has_success and is_react_skip:
            break
    return {"has_success": has_success, "is_react_skip": is_react_skip}


def lookup_url(url: str) -> list[dict]:
    """Find all tracker entries matching this URL (normalized). Returns row details."""
    from hunter import db as _db
    rows = _db.get_by_norm_url(normalize_url(url))
    return [
        {
            "row": 0,
            "company": r["company"],
            "title": r["title"],
            "ats": r["ats"],
            "folder": r["folder"],
            "sent": r["sent"],
        }
        for r in rows
    ]


def lookup_company(company: str) -> list[dict]:
    """Find all tracker entries matching this company (normalized). Returns row details."""
    norm = normalize_company(company)
    if not norm:
        return []
    from hunter import db as _db
    return [
        {
            "row": 0,
            "company": r["Company"],
            "title": r["Job Title"],
            "ats": r["ATS %"],
            "folder": r["Folder"],
            "sent": r["Sent"],
        }
        for r in _db.get_all_rows()
        if normalize_company(r.get("Company", "")) == norm
    ]


def is_known(url: str, company: str = "", title: str = "") -> bool:
    """Check if a job is already in tracker (by URL or company+title)."""
    from hunter import db as _db
    ct = dedup_key(company, title) if company and title else ""
    return _db.is_known(normalize_url(url), ct)


def _db_insert(values: list, *, replace: bool = False) -> None:
    """Insert a values list (parallel to TRACKER_HEADERS) into SQLite. Best-effort."""
    try:
        from hunter import db as _db
        _db.insert_job(dict(zip(TRACKER_HEADERS, (str(v or "") for v in values))), replace=replace)
    except Exception as exc:
        import logging as _log
        _log.getLogger(__name__).warning("[tracker] db insert failed: %s", exc)


def _company_from_content(content: dict) -> str:
    """Prefer explicit company_name; fall back to output folder name."""
    cn = (content.get("company_name") or "").strip()
    if cn:
        return cn
    folder_name = Path(str(content.get("output_folder") or "")).name
    # Remove collision suffixes: Company_2, Company_3...
    s = re.sub(r"_(\d+)$", "", folder_name)
    # Legacy flat structure: Company_YYYY-MM-DD
    m = re.search(r"^(.+)_[0-9]{4}-[0-9]{2}-[0-9]{2}$", s)
    return (m.group(1) if m else s) or "Unknown"


def _parse_ats_score(raw: str) -> tuple[str, int | None]:
    """Normalize ATS score for tracker display and optional color coding."""
    value = (raw or "").strip()
    if not value:
        return "", None

    # Support 10-point scales from LLM output: "8/10", "8.5 / 10".
    m10 = re.search(r"(\d{1,2}(?:[.,]\d+)?)\s*/\s*10\b", value)
    if m10:
        base = float(m10.group(1).replace(",", "."))
        score = max(0, min(int(round(base * 10)), 100))
        return f"{score}%", score

    # Support common LLM variants: "85", "85%", "score: 85/100"
    m = re.search(r"\d{1,3}", value)
    if not m:
        return value, None

    score = max(0, min(int(m.group(0)), 100))
    return f"{score}%", score


def remove_manual_pending_rows(url: str) -> int:
    """Delete tracker rows for this URL where ATS is MANUAL (successful apply supersedes)."""
    from hunter import db as _db
    norm = normalize_url(url)
    db_deleted = _db.delete_where(norm, MANUAL_PENDING_ATS)

    if not TRACKER_PATH.exists():
        return db_deleted
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    deleted = 0
    for row_num in range(ws.max_row, 1, -1):
        row_url = str(ws.cell(row=row_num, column=URL_COL_INDEX).value or "").strip()
        if not row_url or normalize_url(row_url) != norm:
            continue
        ats = str(ws.cell(row=row_num, column=ATS_COL_INDEX).value or "").strip().upper()
        if ats == MANUAL_PENDING_ATS:
            ws.delete_rows(row_num)
            deleted += 1
    if deleted:
        _save_with_retry(wb)
    else:
        wb.close()
    return db_deleted or deleted


def manual_jobleads_job_posting_path(url: str) -> Path | None:
    """Return path to job_posting.txt for a MANUAL JobLeads row, or None."""
    from hunter import db as _db
    norm = normalize_url(url)
    for r in _db.get_by_norm_url(norm):
        if r["ats"].upper() != MANUAL_PENDING_ATS:
            continue
        folder_raw = r["folder"]
        if not folder_raw:
            return None
        p = Path(folder_raw)
        if not p.is_absolute():
            p = TRACKER_PATH.parent / folder_raw
        return p / "job_posting.txt"
    return None


def has_manual_pending(url: str) -> bool:
    """True if tracker already has a MANUAL row for this URL."""
    return any(
        (r.get("ats") or "").strip().upper() == MANUAL_PENDING_ATS
        for r in lookup_url(url)
    )


def get_all_manual_pending() -> list[dict]:
    """Return all MANUAL-pending rows as dicts with keys: url, folder, company, title, row."""
    from hunter import db as _db
    return [
        {"url": r["url"], "folder": r["folder"], "company": r["company"],
         "title": r["title"], "row": r["row"]}
        for r in _db.get_by_ats(MANUAL_PENDING_ATS)
        if r["url"]
    ]


def latest_manual_pending() -> dict[str, str] | None:
    """Return latest MANUAL row info (url, folder) or None."""
    from hunter import db as _db
    rows = [r for r in _db.get_by_ats(MANUAL_PENDING_ATS) if r["url"]]
    if not rows:
        return None
    last = rows[-1]
    return {"url": last["url"], "folder": last["folder"]}


def add_manual_jobleads_pending(
    *,
    url: str,
    company: str,
    title: str,
    folder_abs: Path,
) -> bool:
    """Append MANUAL row for JobLeads when description cannot be fetched.

    Skips when URL already has a success row or any tracker row (dedup / FAIL / MANUAL).
    """
    if has_successful_entry(url) or lookup_url(url):
        return False
    wb, ws = _load_or_create()
    today = date.today().strftime("%Y-%m-%d")
    next_row = ws.max_row + 1
    folder_str = str(folder_abs).replace("\\", "/")
    values = [
        today,
        company.strip() or "Unknown",
        title.strip() or "Unknown",
        "",
        MANUAL_PENDING_ATS,
        url,
        folder_str,
        "",
        "",
        "Paste job text into job_posting.txt in Folder, then re-run apply (same URL).",
        _new_row_id(),
    ]
    _db_insert(values)
    row_font = Font(name="Calibri", size=11)
    manual_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        cell.fill = manual_fill
        if col == URL_COL_INDEX and val:
            cell.hyperlink = str(val)
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    _save_with_retry(wb)
    return True


def add_applied(content: dict, force: bool = False) -> bool:
    """Append a successful apply row. Returns True when a row was written."""
    company = _company_from_content(content)
    job_title = str(content.get("job_title", "") or "")
    stack = str(content.get("stack", "") or "")
    apply_url = str(content.get("apply_url", "") or "")
    if apply_url:
        remove_manual_pending_rows(apply_url)
    folder = str(content.get("output_folder", "") or "")
    to_learn = str(content.get("to_learn", "") or "")
    ats_raw = str(content.get("ats_score", "") or "")
    ats_display, ats_numeric = _parse_ats_score(ats_raw)
    today = date.today().strftime("%Y-%m-%d")
    norm_url = normalize_url(apply_url) if apply_url else ""

    # Keep historical behavior: do not duplicate successful rows unless force mode.
    if norm_url and has_successful_entry(apply_url) and not force:
        return False

    from hunter import db as _db
    is_reapply = bool(norm_url and _db.get_by_norm_url(norm_url))

    wb, ws = _load_or_create()
    next_row = ws.max_row + 1
    row_font = Font(name="Calibri", size=11)
    values = [
        today,
        company,
        job_title,
        stack,
        ats_display,
        apply_url,
        folder,
        "",
        "+" if is_reapply else "",
        to_learn,
        _new_row_id(),
    ]

    _db_insert(values, replace=force)

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        if col == URL_COL_INDEX and val:
            cell.hyperlink = str(val)
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        if col == ATS_COL_INDEX and ats_display:
            cell.alignment = Alignment(horizontal="center")
            if ats_numeric is not None:
                if ats_numeric >= 80:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(name="Calibri", size=11, color="276221", bold=True)
                elif ats_numeric >= 60:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(name="Calibri", size=11, color="9C6500", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(name="Calibri", size=11, color="9C0006", bold=True)
        if col in (SENT_COL_INDEX, 9):
            cell.alignment = Alignment(horizontal="center")

    # Alternate row color for readability.
    if next_row % 2 == 0:
        fill = PatternFill("solid", fgColor="EEF2FA")
        for col in range(1, len(TRACKER_HEADERS) + 1):
            ws.cell(row=next_row, column=col).fill = fill

    _save_with_retry(wb)
    return True


def add_skipped(job: Job) -> dict | None:
    """Append a SKIP row to tracker. Returns the row dict (with ID) or None if already known."""
    if is_known(job.url, job.company, job.title):
        return None
    wb, ws = _load_or_create()
    today = date.today().strftime("%Y-%m-%d")
    next_row = ws.max_row + 1
    row_id = _new_row_id()

    values = [
        today,           # Date
        job.company,     # Company
        job.title,       # Job Title
        "",              # Stack  (unknown at this point)
        "SKIP",          # ATS %  (repurposed for status)
        job.url,         # URL
        "",              # Folder
        "",              # Sent
        "",              # Re-application
        "",              # To Learn
        row_id,          # ID
    ]

    _db_insert(values)
    row_font = Font(name="Calibri", size=11)
    skip_fill = PatternFill("solid", fgColor="D9D9D9")

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        cell.fill = skip_fill
        if col == URL_COL_INDEX:
            cell.hyperlink = job.url
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    _save_with_retry(wb)
    return dict(zip(TRACKER_HEADERS, values))


def is_react_skipped(url: str) -> bool:
    """True if tracker already has a React-skip row (SKIP + Sent='—') for this URL."""
    return get_url_status_flags(url)["is_react_skip"]


def add_react_skipped(content: dict, url: str) -> None:
    """Write a SKIP row for a React-only job. Sent='—' marks it as stack-filtered.

    Light yellow fill distinguishes these rows from geo/tech SKIP rows (grey).
    The URL is recorded so the job is never re-surfaced by the hunter or apply_agent.
    """
    company = (content.get("company_name") or "").strip()
    title   = (content.get("job_title")    or "").strip()
    if is_known(url, company, title):
        return
    wb, ws = _load_or_create()
    today = date.today().strftime("%Y-%m-%d")
    next_row = ws.max_row + 1

    values = [
        today,                          # Date
        company,                        # Company
        title,                          # Job Title
        content.get("stack", ""),       # Stack
        "SKIP",                         # ATS %
        url,                            # URL
        "",                             # Folder
        "—",                            # Sent  ← marks React-skip
        "",                             # Re-application
        "",                             # To Learn
        _new_row_id(),                  # ID
    ]

    _db_insert(values)
    row_font = Font(name="Calibri", size=11)
    row_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        cell.fill = row_fill
        if col == URL_COL_INDEX:
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    _save_with_retry(wb)


def add_expired(url: str, company: str = "", title: str = "") -> None:
    """Write an EXPIRED row — offer was no longer active when fetched.

    Orange fill distinguishes these from SKIP (grey) and React-skip (yellow).
    URL is recorded so the job is never re-fetched.
    """
    if is_known(url, company, title):
        return
    wb, ws = _load_or_create()
    today = date.today().strftime("%Y-%m-%d")
    next_row = ws.max_row + 1

    values = [
        today,           # Date
        company,         # Company
        title,           # Job Title
        "",              # Stack
        "EXPIRED",       # ATS %
        url,             # URL
        "",              # Folder
        "",              # Sent
        "",              # Re-application
        "",              # To Learn
        _new_row_id(),   # ID
    ]

    _db_insert(values)
    row_font = Font(name="Calibri", size=11)
    row_fill = PatternFill("solid", fgColor="FCE4D6")  # light orange

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        cell.fill = row_fill
        if col == URL_COL_INDEX:
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    _save_with_retry(wb)


def add_failed(job: Job) -> None:
    """Append a FAIL row so the job is not retried on next hunt.
    User can delete the row from Excel to retry manually."""
    if is_known(job.url, job.company, job.title):
        return
    wb, ws = _load_or_create()
    today = date.today().strftime("%Y-%m-%d")
    next_row = ws.max_row + 1

    values = [
        today,           # Date
        job.company,     # Company
        job.title,       # Job Title
        "",              # Stack
        "FAIL",          # ATS %  (repurposed for status)
        job.url,         # URL
        "",              # Folder
        "",              # Sent
        "",              # Re-application
        "",              # To Learn
        _new_row_id(),   # ID
    ]

    _db_insert(values)
    row_font = Font(name="Calibri", size=11)
    fail_fill = PatternFill("solid", fgColor="F4CCCC")  # light red

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = row_font
        cell.fill = fail_fill
        if col == URL_COL_INDEX:
            cell.hyperlink = job.url
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    _save_with_retry(wb)


def iter_unsent_rows() -> list[dict]:
    """Return unsent tracker rows (no Sent value, excluding SKIP).

    Each dict has keys: id, date, company, title, stack, ats, url, folder,
    sent, reapp, to_learn, row_num.
    """
    from hunter import db as _db
    return _db.get_unsent_rows()


def apply_sent_updates(updates: dict[str, str]) -> int:
    """Write Sent values (e.g. EXPIRED) back into tracker.xlsx.

    updates: {row_id: sent_value} — only non-empty sent_value entries.
    Returns number of rows updated.
    """
    if not updates:
        return 0
    from hunter import db as _db
    updated = sum(_db.update_sent(rid, sv) for rid, sv in updates.items() if sv)
    # Mirror to Excel for the human-readable view.
    if updated and TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
        for row_num in range(2, ws.max_row + 1):
            row_id = str(ws.cell(row=row_num, column=ID_COL_INDEX).value or "").strip()
            if row_id and row_id in updates and updates[row_id]:
                ws.cell(row=row_num, column=SENT_COL_INDEX).value = updates[row_id]
        _save_with_retry(wb)
    return updated


_REAPP_COL_INDEX = 9    # "Re-application"
_TO_LEARN_COL_INDEX = 10  # "To Learn"


def apply_pull_updates(rows: list[dict]) -> int:
    """Write Sheets-sourced field changes back to tracker.xlsx (pull sync).

    rows: list of full row dicts (with 'ID') where Sheets had a newer value.
    Updates Sent, Re-application, To Learn columns. Returns count of rows updated.
    """
    if not rows:
        return 0
    from hunter import db as _db
    updated = 0
    for row_dict in rows:
        row_id = row_dict.get("ID", "").strip()
        if not row_id:
            continue
        updated += _db.update_user_fields(
            row_id,
            sent=row_dict.get("Sent", ""),
            reapply=row_dict.get("Re-application", ""),
            to_learn=row_dict.get("To Learn", ""),
        )
    # Mirror to Excel for the human-readable view.
    if updated and TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        ws = wb.active
        for row_num in range(2, ws.max_row + 1):
            row_id = str(ws.cell(row=row_num, column=ID_COL_INDEX).value or "").strip()
            if not row_id:
                continue
            for row_dict in rows:
                if row_dict.get("ID", "").strip() != row_id:
                    continue
                ws.cell(row=row_num, column=SENT_COL_INDEX).value = row_dict.get("Sent", "")
                ws.cell(row=row_num, column=_REAPP_COL_INDEX).value = row_dict.get("Re-application", "")
                ws.cell(row=row_num, column=_TO_LEARN_COL_INDEX).value = row_dict.get("To Learn", "")
                break
        _save_with_retry(wb)
    return updated


def get_folder_by_url(url: str) -> str | None:
    """Return the Folder value for a given job URL, or None if not found.

    Normalizes the URL before comparing (strip trailing slash, lowercase scheme).
    Returns the raw string from the Folder column (e.g. 'Applications/2026-05-11/PeopleMore_3').
    """
    target = normalize_url(url)
    if not target:
        return None
    from hunter import db as _db
    for r in _db.get_by_norm_url(target):
        if r.get("folder"):
            return r["folder"]
    return None


# gsheets COLUMNS order: Date, Company, Job Title, Stack, ATS %, URL,
#                        Folder, Sent, Re-application, To Learn, ID
_GSHEETS_COLS = [
    "Date", "Company", "Job Title", "Stack", "ATS %", "URL",
    "Folder", "Sent", "Re-application", "To Learn", "ID",
]

def read_all_tracker_rows() -> list[dict]:
    """Read every data row as a dict keyed by TRACKER_HEADERS names.

    Rows with no ID are skipped. Used by gsheets_sync.push_missing_rows().
    """
    from hunter import db as _db
    return _db.get_all_rows()


def export_to_excel(path: Path = TRACKER_PATH) -> int:
    """Regenerate tracker.xlsx from SQLite. Overwrites the file.

    Returns number of rows written. Uses the same formatting as normal writes.
    Called by /export Telegram command.
    """
    from hunter import db as _db
    rows = _db.get_all_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    header_fill = PatternFill("solid", fgColor="2B579A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    widths = [12, 20, 30, 12, 8, 50, 40, 8, 16, 35, 12]

    for col, (header, width) in enumerate(zip(TRACKER_HEADERS, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    _ATS_STATUS_FILLS = {
        "SKIP":    PatternFill("solid", fgColor="D9D9D9"),
        "FAIL":    PatternFill("solid", fgColor="F4CCCC"),
        "EXPIRED": PatternFill("solid", fgColor="FCE4D6"),
        "MANUAL":  PatternFill("solid", fgColor="FFF2CC"),
    }

    for row_num, row in enumerate(rows, 2):
        ats = str(row.get("ATS %") or "").strip()
        sent = str(row.get("Sent") or "").strip()

        # Determine row fill
        if ats == "SKIP" and sent in REACT_SKIP_SENT_MARKERS:
            row_fill = PatternFill("solid", fgColor="FFF2CC")  # react-skip: light yellow
        elif ats in _ATS_STATUS_FILLS:
            row_fill = _ATS_STATUS_FILLS[ats]
        elif row_num % 2 == 0:
            row_fill = PatternFill("solid", fgColor="EEF2FA")
        else:
            row_fill = None

        _, ats_numeric = _parse_ats_score(ats)

        for col, header in enumerate(TRACKER_HEADERS, 1):
            val = row.get(header, "")
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name="Calibri", size=11)
            if row_fill:
                cell.fill = row_fill
            if col == URL_COL_INDEX and val:
                cell.hyperlink = str(val)
                cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
            if col == ATS_COL_INDEX and ats and ats_numeric is not None:
                cell.alignment = Alignment(horizontal="center")
                if ats_numeric >= 80:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(name="Calibri", size=11, color="276221", bold=True)
                elif ats_numeric >= 60:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(name="Calibri", size=11, color="9C6500", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(name="Calibri", size=11, color="9C0006", bold=True)
            if col in (SENT_COL_INDEX, 9):
                cell.alignment = Alignment(horizontal="center")

    _save_with_retry(wb, retries=3, delay=2.0)
    return len(rows)
