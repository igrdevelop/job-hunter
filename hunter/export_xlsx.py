"""
hunter/export_xlsx.py — Generate tracker.xlsx from SQLite on demand.

Used by the /export Telegram command (step 5.3).
openpyxl is ONLY needed here and in generate_docs.py — not in tracker.py.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from hunter.tracker import TRACKER_HEADERS, read_all_tracker_rows

logger = logging.getLogger(__name__)

# Header row style — mirrors the original tracker formatting
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="center", wrap_text=True)

# Approximate column widths (characters) — nice default for human reading
_COL_WIDTHS = {
    "Date": 12,
    "Company": 22,
    "Job Title": 30,
    "Stack": 18,
    "ATS %": 8,
    "URL": 40,
    "Folder": 35,
    "Sent": 12,
    "Re-application": 14,
    "To Learn": 20,
    "ID": 10,
    "Drive URL": 40,
    "Confirmation": 14,
    "Answer": 20,
}


def export_tracker_xlsx(output_path: Path) -> int:
    """Write all tracker rows to *output_path* as a styled xlsx workbook.

    Returns the number of data rows written.
    Does NOT overwrite the live tracker.xlsx — callers should use a temp path
    or an explicit destination that does not conflict with the DB-backed file.
    """
    rows = read_all_tracker_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracker"

    # Header row
    ws.append(TRACKER_HEADERS)
    for col_idx, header in enumerate(TRACKER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(header, 15)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Data rows
    for row_dict in rows:
        ws.append([row_dict.get(h, "") for h in TRACKER_HEADERS])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[export] wrote %d rows to %s", len(rows), output_path)
    return len(rows)
