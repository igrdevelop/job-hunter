"""
to_send.py — manage to_send.xlsx, a human-editable view of unsent applications.

Workflow:
  1. Bot writes to tracker.xlsx (canonical store).
  2. sync_and_rebuild() reads "Sent" marks from to_send.xlsx, copies them
     back to tracker.xlsx, then rebuilds to_send.xlsx from tracker rows
     that still have an empty Sent column (SKIP rows are excluded).
  3. User opens to_send.xlsx in Excel, fills in "Sent" column (any value,
     e.g. a date, "+" or "ok"), saves and closes.
  4. Next sync (on new apply or /sync_sent Telegram command) picks up the
     changes and removes those rows from to_send.xlsx.

to_send.xlsx is safe to have open; if it is locked the rebuild step
logs a warning and returns without raising (main apply flow is unaffected).
"""

import logging
import time

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hunter.config import TO_SEND_PATH
from hunter.tracker import (
    ID_COL_INDEX,
    SENT_COL_INDEX,
    URL_COL_INDEX,
    apply_sent_updates,
    iter_rows_for_to_send,
)

logger = logging.getLogger(__name__)

TO_SEND_HEADERS = [
    "Date", "Company", "Job Title", "Stack",
    "ATS %", "URL", "Folder", "Sent", "Re-application", "To Learn", "ID",
]
_WIDTHS = [12, 22, 32, 14, 8, 50, 40, 12, 14, 35, 10]

# ATS fill colours to visually match tracker.xlsx
_ATS_FILLS = {
    "FAIL":   PatternFill("solid", fgColor="F4CCCC"),
    "MANUAL": PatternFill("solid", fgColor="FFF2CC"),
}


def _save_to_send_safe(wb: openpyxl.Workbook, retries: int = 3, delay: float = 2.0) -> bool:
    """Save to_send.xlsx; on PermissionError (file open) log warning and return False."""
    for attempt in range(1, retries + 1):
        try:
            wb.save(TO_SEND_PATH)
            return True
        except PermissionError:
            if attempt == retries:
                logger.warning(
                    "[to_send] to_send.xlsx is open in Excel — rebuild skipped. "
                    "Close the file and run /sync_sent again."
                )
                return False
            logger.debug("[to_send] File locked, retry %d/%d in %.0fs…", attempt, retries, delay)
            time.sleep(delay)
    return False


def _write_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F5C99")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    for col, (header, width) in enumerate(zip(TO_SEND_HEADERS, _WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20


def _write_data_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    r: dict,
) -> None:
    row_font = Font(name="Calibri", size=11)
    ats = r["ats"]
    row_fill = _ATS_FILLS.get(ats)

    values = [
        r["date"],
        r["company"],
        r["title"],
        r["stack"],
        ats,
        r["url"],
        r["folder"],
        r["sent"],
        r["reapp"],
        r["to_learn"],
        r["id"],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = row_font
        if row_fill:
            cell.fill = row_fill
        if col == URL_COL_INDEX and val:
            cell.hyperlink = str(val)
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        if col == SENT_COL_INDEX:
            cell.alignment = Alignment(horizontal="center")

    # Alternate row shade for readability (only for non-status rows)
    if not row_fill and row_idx % 2 == 0:
        alt_fill = PatternFill("solid", fgColor="EEF2FA")
        for col in range(1, len(TO_SEND_HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = alt_fill


# ── Public API ────────────────────────────────────────────────────────────────

def is_excel_open() -> bool:
    """Return True if Excel currently has to_send.xlsx open (lock file exists)."""
    lock = TO_SEND_PATH.parent / f"~${TO_SEND_PATH.name}"
    return lock.exists()


def read_sent_marks() -> dict[str, str]:
    """Read to_send.xlsx and return {row_id: sent_value} for rows with a non-empty Sent."""
    if not TO_SEND_PATH.exists():
        return {}

    if is_excel_open():
        logger.warning(
            "[to_send] to_send.xlsx appears open in Excel — unsaved changes will NOT be read. "
            "Save and close the file, then run /sync_sent again."
        )

    try:
        wb = openpyxl.load_workbook(TO_SEND_PATH, read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("[to_send] Could not open to_send.xlsx: %s", exc)
        return {}

    ws = wb.active
    marks: dict[str, str] = {}
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            row = tuple(row) + ("",) * max(0, ID_COL_INDEX - len(row))
            sent = str(row[SENT_COL_INDEX - 1] or "").strip()
            row_id = str(row[ID_COL_INDEX - 1] or "").strip()
            if sent and row_id:
                marks[row_id] = sent
    finally:
        wb.close()
    return marks


def rebuild() -> bool:
    """Rebuild to_send.xlsx from current tracker rows that have no Sent value.

    Returns True on success, False if the file was locked.
    """
    rows = iter_rows_for_to_send()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ToSend"

    _write_header(ws)

    for i, r in enumerate(rows, 2):
        _write_data_row(ws, i, r)

    ws.freeze_panes = "A2"
    # Hide the ID column — it's an internal key, not meant for manual editing
    ws.column_dimensions[get_column_letter(ID_COL_INDEX)].hidden = True

    return _save_to_send_safe(wb)


def sync_and_rebuild() -> dict:
    """Full sync cycle:
      1. Read Sent marks from to_send.xlsx.
      2. Write them back to tracker.xlsx.
      3. Rebuild to_send.xlsx (only unsent rows remain).

    Returns {"synced": N, "rebuilt": bool}.
    """
    marks = read_sent_marks()
    synced = apply_sent_updates(marks) if marks else 0
    if synced:
        logger.info("[to_send] Synced %d Sent mark(s) → tracker.xlsx", synced)
    rebuilt = rebuild()
    return {"synced": synced, "rebuilt": rebuilt}
