"""
tests/test_db.py — Unit tests for hunter/db.py (SQLite schema, init, migration).
"""

import sqlite3
from pathlib import Path

import pytest

from hunter.db import get_db, init_db, migrate_from_excel, row_to_dict


# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    """Isolated SQLite database path (no auto-migration from real tracker.xlsx)."""
    p = tmp_path / "tracker.db"
    # Pass a non-existent xlsx_path to prevent auto-migration from real tracker.xlsx.
    init_db(p, xlsx_path=tmp_path / "no_tracker.xlsx")
    return p


# ── Schema / init ─────────────────────────────────────────────────────────────


def test_init_db_creates_file(tmp_path: Path) -> None:
    p = tmp_path / "tracker.db"
    assert not p.exists()
    init_db(p)
    assert p.exists()


def test_init_db_creates_applications_table(db_path: Path) -> None:
    with get_db(db_path) as conn:
        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    assert "applications" in tables


def test_init_db_creates_subsystem_health_table(db_path: Path) -> None:
    """Backs hunter.best_effort — see tests/test_best_effort.py for behavior."""
    with get_db(db_path) as conn:
        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    assert "subsystem_health" in tables


def test_init_db_has_all_columns(db_path: Path) -> None:
    expected = {
        "id",
        "date",
        "company",
        "title",
        "stack",
        "ats_status",
        "url",
        "url_norm",
        "folder",
        "sent",
        "reapplication",
        "to_learn",
        "drive_url",
        "confirmation",
        "answer",
        "sheets_row",
        "sheets_dirty",
        "fail_count",
        "cost_usd",
        "ats_verdict",
    }
    with get_db(db_path) as conn:
        cols = {row[1] for row in conn.execute("PRAGMA table_info(applications)")}
    assert expected == cols


def test_init_db_wal_mode(db_path: Path) -> None:
    with get_db(db_path) as conn:
        mode = conn.execute("PRAGMA journal_mode").fetchone()[0]
    assert mode == "wal"


def test_init_db_idempotent(tmp_path: Path) -> None:
    """Calling init_db twice should not raise or corrupt data."""
    p = tmp_path / "tracker.db"
    no_xlsx = tmp_path / "no_tracker.xlsx"
    init_db(p, xlsx_path=no_xlsx)
    # Insert a row
    with get_db(p) as conn:
        conn.execute(
            "INSERT INTO applications (id, company, title) VALUES ('abc12345','Acme','Dev')"
        )
    # Second init should be a no-op (DB already exists)
    init_db(p, xlsx_path=no_xlsx)
    with get_db(p) as conn:
        count = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    assert count == 1


def test_init_db_creates_indexes(db_path: Path) -> None:
    with get_db(db_path) as conn:
        indexes = {
            r[1]
            for r in conn.execute(
                "SELECT * FROM sqlite_master WHERE type='index' AND tbl_name='applications'"
            )
        }
    assert "idx_url_norm" in indexes
    assert "idx_ats" in indexes


# ── get_db context manager ───────────────────────────────────────────────────


def test_get_db_commits_on_success(db_path: Path) -> None:
    with get_db(db_path) as conn:
        conn.execute("INSERT INTO applications (id, company, title) VALUES ('aaa11111','X','Y')")
    # Open a fresh connection — data should be persisted
    with get_db(db_path) as conn:
        row = conn.execute("SELECT company FROM applications WHERE id='aaa11111'").fetchone()
    assert row is not None
    assert row["company"] == "X"


def test_get_db_rolls_back_on_exception(db_path: Path) -> None:
    try:
        with get_db(db_path) as conn:
            conn.execute(
                "INSERT INTO applications (id, company, title) VALUES ('bbb22222','Y','Z')"
            )
            raise RuntimeError("simulated failure")
    except RuntimeError:
        pass
    with get_db(db_path) as conn:
        row = conn.execute("SELECT * FROM applications WHERE id='bbb22222'").fetchone()
    assert row is None


def test_get_db_row_factory(db_path: Path) -> None:
    with get_db(db_path) as conn:
        conn.execute(
            "INSERT INTO applications (id, company, title, url_norm) "
            "VALUES ('ccc33333','Foo','Bar','https://foo.com')"
        )
    with get_db(db_path) as conn:
        row = conn.execute("SELECT * FROM applications WHERE id='ccc33333'").fetchone()
    assert isinstance(row, sqlite3.Row)
    assert row["company"] == "Foo"
    assert row["title"] == "Bar"


# ── row_to_dict ───────────────────────────────────────────────────────────────


def test_row_to_dict(db_path: Path) -> None:
    with get_db(db_path) as conn:
        conn.execute(
            "INSERT INTO applications (id, company, title) VALUES ('ddd44444','Baz','Qux')"
        )
    with get_db(db_path) as conn:
        row = conn.execute("SELECT * FROM applications WHERE id='ddd44444'").fetchone()
    d = row_to_dict(row)
    assert isinstance(d, dict)
    assert d["company"] == "Baz"
    assert d["title"] == "Qux"
    assert "sheets_dirty" in d


# ── migrate_from_excel ────────────────────────────────────────────────────────


def _make_xlsx(path: Path) -> None:
    """Create a minimal tracker.xlsx with two data rows."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Date",
        "Company",
        "Job Title",
        "Stack",
        "ATS %",
        "URL",
        "Folder",
        "Sent",
        "Re-application",
        "To Learn",
        "ID",
        "Drive URL",
        "Confirmation",
        "Answer",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    # Row 1 — applied with ATS score
    ws.cell(row=2, column=1, value="2026-05-01")
    ws.cell(row=2, column=2, value="Acme")
    ws.cell(row=2, column=3, value="Frontend Dev")
    ws.cell(row=2, column=4, value="Angular")
    ws.cell(row=2, column=5, value="82%")
    ws.cell(row=2, column=6, value="https://acme.com/jobs/42")
    ws.cell(row=2, column=7, value="Applications/2026-05-01/Acme")
    ws.cell(row=2, column=11, value="abcd1234")

    # Row 2 — skipped
    ws.cell(row=3, column=1, value="2026-05-02")
    ws.cell(row=3, column=2, value="Beta Corp")
    ws.cell(row=3, column=3, value="Angular Developer")
    ws.cell(row=3, column=5, value="SKIP")
    ws.cell(row=3, column=6, value="https://beta.com/jobs/7")
    ws.cell(row=3, column=11, value="efgh5678")

    # Row with no ID — must be skipped by migration
    ws.cell(row=4, column=2, value="No ID Corp")
    ws.cell(row=4, column=6, value="https://noid.com/jobs/1")

    wb.save(path)


def test_migrate_creates_rows(tmp_path: Path) -> None:
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)
    init_db(db, xlsx_path=tmp_path / "no_tracker.xlsx")  # no auto-migrate
    n = migrate_from_excel(xlsx, db)
    assert n == 2
    with get_db(db) as conn:
        count = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    assert count == 2


def test_migrate_skips_no_id_row(tmp_path: Path) -> None:
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)
    init_db(db)
    migrate_from_excel(xlsx, db)
    with get_db(db) as conn:
        # No-ID row must not exist
        row = conn.execute("SELECT * FROM applications WHERE company='No ID Corp'").fetchone()
    assert row is None


def test_migrate_stores_url_norm(tmp_path: Path) -> None:
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)
    init_db(db)
    migrate_from_excel(xlsx, db)
    with get_db(db) as conn:
        row = conn.execute("SELECT url_norm FROM applications WHERE id='abcd1234'").fetchone()
    assert row is not None
    # normalize_url strips trailing slash / downcases host
    assert "acme.com" in row["url_norm"]


def test_migrate_is_idempotent(tmp_path: Path) -> None:
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)
    init_db(db, xlsx_path=tmp_path / "no_tracker.xlsx")  # no auto-migrate
    n1 = migrate_from_excel(xlsx, db)
    n2 = migrate_from_excel(xlsx, db)
    assert n1 == 2
    assert n2 == 0  # all rows already present
    with get_db(db) as conn:
        count = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    assert count == 2


def test_migrate_missing_xlsx_returns_zero(tmp_path: Path) -> None:
    db = tmp_path / "tracker.db"
    init_db(db)
    n = migrate_from_excel(tmp_path / "nonexistent.xlsx", db)
    assert n == 0


def test_init_db_auto_migrates_when_xlsx_present(tmp_path: Path) -> None:
    """init_db should auto-migrate when xlsx_path is given and DB does not exist yet."""
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)

    init_db(db, xlsx_path=xlsx)  # explicit xlsx → will migrate

    with get_db(db) as conn:
        count = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    assert count == 2


def test_init_db_no_auto_migrate_when_db_exists(tmp_path: Path) -> None:
    """init_db should NOT re-migrate if DB already exists."""
    xlsx = tmp_path / "tracker.xlsx"
    db = tmp_path / "tracker.db"
    _make_xlsx(xlsx)

    # First init — migrates
    init_db(db, xlsx_path=xlsx)
    # Second init — DB already exists, migration skipped
    init_db(db, xlsx_path=xlsx)

    with get_db(db) as conn:
        count = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    assert count == 2


# ── Sheets columns ────────────────────────────────────────────────────────────


def test_sheets_dirty_default_is_zero(db_path: Path) -> None:
    with get_db(db_path) as conn:
        conn.execute("INSERT INTO applications (id, company, title) VALUES ('zzz99999','X','Y')")
    with get_db(db_path) as conn:
        row = conn.execute(
            "SELECT sheets_dirty, sheets_row FROM applications WHERE id='zzz99999'"
        ).fetchone()
    assert row["sheets_dirty"] == 0
    assert row["sheets_row"] is None


def test_can_set_sheets_row_and_dirty(db_path: Path) -> None:
    with get_db(db_path) as conn:
        conn.execute("INSERT INTO applications (id, company, title) VALUES ('www88888','A','B')")
        conn.execute("UPDATE applications SET sheets_row=5, sheets_dirty=1 WHERE id='www88888'")
    with get_db(db_path) as conn:
        row = conn.execute(
            "SELECT sheets_row, sheets_dirty FROM applications WHERE id='www88888'"
        ).fetchone()
    assert row["sheets_row"] == 5
    assert row["sheets_dirty"] == 1
