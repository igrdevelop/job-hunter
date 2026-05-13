"""Tests for hunter/to_send.py sync and rebuild logic."""

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

import hunter.tracker as tracker
import hunter.to_send as to_send
from hunter.models import Job


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def patch_paths(tmp_path, monkeypatch):
    tracker_path = tmp_path / "tracker.xlsx"
    send_path = tmp_path / "to_send.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)
    monkeypatch.setattr(to_send, "TO_SEND_PATH", send_path)
    # Also patch the config import inside tracker_service if used
    import hunter.config as cfg
    monkeypatch.setattr(cfg, "TRACKER_PATH", tracker_path)
    monkeypatch.setattr(cfg, "TO_SEND_PATH", send_path)
    return tracker_path, send_path


def _make_job(n: int = 1) -> Job:
    return Job(
        title=f"Frontend Dev {n}",
        company=f"Corp{n}",
        location="remote",
        salary=None,
        url=f"https://example.com/job/{n}",
        source="test",
    )


def _applied_content(n: int = 1) -> dict:
    return {
        "company_name": f"Corp{n}",
        "job_title": f"Frontend Dev {n}",
        "stack": "Angular",
        "apply_url": f"https://example.com/job/{n}",
        "output_folder": f"Applications/2026-01-01/Corp{n}",
        "ats_score": "85%",
        "to_learn": "",
    }


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_rebuild_contains_only_unsent_rows(patch_paths):
    """rebuild() puts only rows with empty Sent into to_send.xlsx."""
    tracker.add_applied(_applied_content(1))
    tracker.add_applied(_applied_content(2))

    # Mark row 1 as sent directly in tracker
    wb = openpyxl.load_workbook(patch_paths[0])
    ws = wb.active
    ws.cell(row=2, column=tracker.SENT_COL_INDEX).value = "2026-04-21"
    wb.save(patch_paths[0])

    to_send.rebuild()

    wb2 = openpyxl.load_workbook(patch_paths[1])
    ws2 = wb2.active
    companies = [ws2.cell(row=r, column=2).value for r in range(2, ws2.max_row + 1)]
    assert "Corp1" not in companies
    assert "Corp2" in companies


def test_rebuild_excludes_pure_skip_rows(patch_paths):
    """SKIP rows are never shown in to_send.xlsx."""
    job = _make_job(3)
    tracker.add_skipped(job)
    to_send.rebuild()

    if patch_paths[1].exists():
        wb = openpyxl.load_workbook(patch_paths[1])
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        companies = [r[1] for r in rows if r[1]]
        assert "Corp3" not in companies


def test_rebuild_includes_fail_and_manual_rows(patch_paths):
    """FAIL and MANUAL rows should appear in to_send.xlsx (user may want to act)."""
    job_fail = _make_job(4)
    tracker.add_failed(job_fail)

    tracker.add_manual_jobleads_pending(
        url="https://example.com/job/5",
        company="Corp5",
        title="Frontend Dev 5",
        folder_abs=patch_paths[0].parent / "Applications" / "Corp5",
    )

    to_send.rebuild()

    wb = openpyxl.load_workbook(patch_paths[1])
    ws = wb.active
    companies = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "Corp4" in companies
    assert "Corp5" in companies


def test_sync_writes_sent_value_to_tracker(patch_paths):
    """Sent value written in to_send.xlsx is copied as-is into tracker.xlsx."""
    tracker.add_applied(_applied_content(6))

    # Get the ID assigned to row 2
    wb = openpyxl.load_workbook(patch_paths[0])
    ws = wb.active
    row_id = ws.cell(row=2, column=tracker.ID_COL_INDEX).value
    wb.close()

    # Simulate user filling in Sent in to_send.xlsx
    to_send.rebuild()
    wb2 = openpyxl.load_workbook(patch_paths[1])
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=2):
        id_cell = row[tracker.ID_COL_INDEX - 1]
        if str(id_cell.value or "") == str(row_id):
            row[tracker.SENT_COL_INDEX - 1].value = "2026-04-21"
    wb2.save(patch_paths[1])

    result = to_send.sync_and_rebuild()
    assert result["synced"] == 1

    # Verify tracker has the Sent value
    wb3 = openpyxl.load_workbook(patch_paths[0])
    ws3 = wb3.active
    sent_val = ws3.cell(row=2, column=tracker.SENT_COL_INDEX).value
    assert str(sent_val) == "2026-04-21"


def test_sent_value_copied_as_is(patch_paths):
    """Arbitrary sent values like '+' or 'ok' are copied verbatim."""
    tracker.add_applied(_applied_content(7))

    wb = openpyxl.load_workbook(patch_paths[0])
    ws = wb.active
    row_id = ws.cell(row=2, column=tracker.ID_COL_INDEX).value
    wb.close()

    to_send.rebuild()
    wb2 = openpyxl.load_workbook(patch_paths[1])
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=2):
        if str(row[tracker.ID_COL_INDEX - 1].value or "") == str(row_id):
            row[tracker.SENT_COL_INDEX - 1].value = "+"
    wb2.save(patch_paths[1])

    to_send.sync_and_rebuild()

    wb3 = openpyxl.load_workbook(patch_paths[0])
    ws3 = wb3.active
    assert ws3.cell(row=2, column=tracker.SENT_COL_INDEX).value == "+"


def test_sent_row_disappears_from_to_send_after_sync(patch_paths):
    """After sync, a sent row is no longer in to_send.xlsx."""
    tracker.add_applied(_applied_content(8))

    wb = openpyxl.load_workbook(patch_paths[0])
    ws = wb.active
    row_id = ws.cell(row=2, column=tracker.ID_COL_INDEX).value
    wb.close()

    to_send.rebuild()
    wb2 = openpyxl.load_workbook(patch_paths[1])
    ws2 = wb2.active
    for row in ws2.iter_rows(min_row=2):
        if str(row[tracker.ID_COL_INDEX - 1].value or "") == str(row_id):
            row[tracker.SENT_COL_INDEX - 1].value = "ok"
    wb2.save(patch_paths[1])

    to_send.sync_and_rebuild()

    wb3 = openpyxl.load_workbook(patch_paths[1])
    ws3 = wb3.active
    ids_in_send = [
        ws3.cell(row=r, column=tracker.ID_COL_INDEX).value
        for r in range(2, ws3.max_row + 1)
    ]
    assert row_id not in ids_in_send


def test_missing_to_send_file_noop(patch_paths):
    """sync_and_rebuild() works fine when to_send.xlsx does not exist yet."""
    tracker.add_applied(_applied_content(9))
    assert not patch_paths[1].exists()
    result = to_send.sync_and_rebuild()
    assert result["synced"] == 0
    assert result["rebuilt"] is True
    assert patch_paths[1].exists()


def test_locked_to_send_logs_warning_not_raises(patch_paths, caplog):
    """PermissionError on save logs a warning but does not propagate."""
    import logging
    tracker.add_applied(_applied_content(10))

    with patch.object(openpyxl.Workbook, "save", side_effect=PermissionError("locked")):
        with caplog.at_level(logging.WARNING, logger="hunter.to_send"):
            result = to_send.rebuild()

    assert result is False
    assert any("locked" in r.message.lower() or "open" in r.message.lower()
               for r in caplog.records)


def test_rebuild_skips_when_libreoffice_lock_present(patch_paths):
    """LibreOffice Calc lock file prevents overwriting to_send.xlsx on disk."""
    tracker.add_applied(_applied_content(11))
    assert to_send.rebuild() is True

    tracker.add_applied(_applied_content(12))
    send_path = patch_paths[1]
    lo_lock = send_path.parent / f".~lock.{send_path.name}#"
    lo_lock.write_text("lock", encoding="utf-8")
    try:
        assert to_send.rebuild() is False
        wb = openpyxl.load_workbook(send_path)
        ws = wb.active
        companies = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "Corp11" in companies
        assert "Corp12" not in companies
    finally:
        lo_lock.unlink(missing_ok=True)


def test_existing_tracker_gets_ids_on_migration(patch_paths):
    """Tracker rows without ID column get IDs assigned on next _load_or_create call."""
    # Build a minimal tracker with no ID column (old format)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    for col, h in enumerate(tracker.TRACKER_HEADERS[:10], 1):
        ws.cell(row=1, column=col, value=h)
    # Write one row with no ID
    ws.cell(row=2, column=1).value = "2026-01-01"
    ws.cell(row=2, column=2).value = "OldCorp"
    ws.cell(row=2, column=tracker.URL_COL_INDEX).value = "https://example.com/old"
    wb.save(patch_paths[0])

    # Trigger migration via _load_or_create
    _, ws2 = tracker._load_or_create()
    row_id = ws2.cell(row=2, column=tracker.ID_COL_INDEX).value
    assert row_id is not None and len(str(row_id)) == 8
