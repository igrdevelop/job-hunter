"""Tests for hunter/export_xlsx.py and the /export Telegram command handler."""

import asyncio
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl

from hunter.export_xlsx import export_tracker_xlsx
from hunter.tracker import TRACKER_HEADERS, normalize_url
from hunter.db import get_db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _insert(
    tracker_db,
    *,
    company: str,
    title: str,
    ats: str = "85%",
    url: str = "",
    date: str = "2026-05-27",
) -> str:
    row_id = uuid.uuid4().hex[:8]
    if not url:
        url = f"https://example.com/{row_id}"
    norm = normalize_url(url)
    with get_db(tracker_db) as conn:
        conn.execute(
            """
            INSERT INTO applications
            (id, date, company, title, ats_status, url, url_norm)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (row_id, date, company, title, ats, url, norm),
        )
    return row_id


def run(coro):
    return asyncio.run(coro)


# ---------------------------------------------------------------------------
# export_tracker_xlsx
# ---------------------------------------------------------------------------


def test_export_creates_file(tmp_path, tracker_db):
    _insert(tracker_db, company="Acme", title="Angular Dev")
    out = tmp_path / "export.xlsx"
    n = export_tracker_xlsx(out)
    assert out.exists()
    assert n == 1


def test_export_empty_db(tmp_path, tracker_db):
    out = tmp_path / "export.xlsx"
    n = export_tracker_xlsx(out)
    assert out.exists()
    assert n == 0


def test_export_header_row(tmp_path, tracker_db):
    out = tmp_path / "export.xlsx"
    export_tracker_xlsx(out)
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=i + 1).value for i in range(len(TRACKER_HEADERS))]
    assert headers == TRACKER_HEADERS
    wb.close()


def test_export_data_row_values(tmp_path, tracker_db):
    row_id = _insert(
        tracker_db,
        company="NASK",
        title="Senior Frontend Developer",
        ats="91%",
        url="https://nask.pl/job/1",
        date="2026-05-21",
    )
    out = tmp_path / "export.xlsx"
    export_tracker_xlsx(out)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    # Row 1 = headers, Row 2 = first data row
    row2 = {
        ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value
        for c in range(1, len(TRACKER_HEADERS) + 1)
    }
    assert row2["Company"] == "NASK"
    assert row2["Job Title"] == "Senior Frontend Developer"
    assert row2["ATS %"] == "91%"
    assert row2["URL"] == "https://nask.pl/job/1"
    assert row2["Date"] == "2026-05-21"
    assert row2["ID"] == row_id
    wb.close()


def test_export_multiple_rows_count(tmp_path, tracker_db):
    for i in range(5):
        _insert(tracker_db, company=f"Company{i}", title=f"Dev {i}")
    out = tmp_path / "export.xlsx"
    n = export_tracker_xlsx(out)
    assert n == 5
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in data_rows if any(r)]
    assert len(data_rows) == 5
    wb.close()


def test_export_creates_parent_dirs(tmp_path, tracker_db):
    out = tmp_path / "deep" / "nested" / "export.xlsx"
    export_tracker_xlsx(out)
    assert out.exists()


# ---------------------------------------------------------------------------
# cmd_export — Telegram handler
# ---------------------------------------------------------------------------


def test_cmd_export_sends_document(tracker_db):
    _insert(tracker_db, company="Acme", title="Dev")
    _insert(tracker_db, company="Beta", title="Dev")

    update = MagicMock()
    context = MagicMock()
    status_msg = AsyncMock()
    update.message.reply_text = AsyncMock(return_value=status_msg)
    update.message.reply_document = AsyncMock()

    async def _run():
        from hunter.commands.export import cmd_export

        await cmd_export(update, context)

    run(_run())

    update.message.reply_text.assert_called_once()
    status_msg.edit_text.assert_called_once()
    edit_call_text = status_msg.edit_text.call_args[0][0]
    assert "2" in edit_call_text  # 2 rows

    update.message.reply_document.assert_called_once()
    doc_kwargs = update.message.reply_document.call_args[1]
    assert doc_kwargs["filename"] == "tracker.xlsx"
    assert "2 rows" in doc_kwargs["caption"]


def test_cmd_export_handles_error(tracker_db):
    update = MagicMock()
    context = MagicMock()
    status_msg = AsyncMock()
    update.message.reply_text = AsyncMock(return_value=status_msg)
    update.message.reply_document = AsyncMock()

    async def _run():
        with patch("hunter.commands.export._do_export", side_effect=RuntimeError("disk full")):
            from hunter.commands.export import cmd_export

            await cmd_export(update, context)

    run(_run())

    status_msg.edit_text.assert_called_once()
    error_text = status_msg.edit_text.call_args[0][0]
    assert "❌" in error_text
    assert "disk full" in error_text
    update.message.reply_document.assert_not_called()
