from pathlib import Path

import openpyxl

from hunter import tracker


def _build_content(url: str, output_folder: Path) -> dict:
    return {
        "company_name": "Acme",
        "job_title": "Senior Frontend Developer",
        "stack": "Angular",
        "ats_score": "85",
        "apply_url": url,
        "output_folder": str(output_folder),
        "to_learn": "State management",
    }


def test_add_applied_writes_success_row(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    content = _build_content(
        "https://example.com/jobs/1?utm_source=mail",
        tmp_path / "Applications" / "2026-04-16" / "Acme",
    )
    written = tracker.add_applied(content, force=False)

    assert written is True
    assert tracker.has_successful_entry("https://example.com/jobs/1")

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 1
    assert rows[0][1] == "Acme"
    assert rows[0][2] == "Senior Frontend Developer"
    assert rows[0][4] == "85%"


def test_add_applied_skips_duplicate_success_when_not_forced(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    content = _build_content(
        "https://example.com/jobs/2?utm_source=mail",
        tmp_path / "Applications" / "2026-04-16" / "Acme",
    )
    assert tracker.add_applied(content, force=False) is True
    assert tracker.add_applied(content, force=False) is False

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    assert len(rows) == 1


def test_add_applied_marks_reapplication_when_forced(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    content = _build_content(
        "https://example.com/jobs/3",
        tmp_path / "Applications" / "2026-04-16" / "Acme",
    )
    assert tracker.add_applied(content, force=False) is True
    assert tracker.add_applied(content, force=True) is True

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    assert len(rows) == 2
    # Re-application column should be marked on the second row.
    assert rows[1][8] == "+"


def test_add_applied_accepts_non_numeric_ats_score(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    content = _build_content(
        "https://example.com/jobs/4",
        tmp_path / "Applications" / "2026-04-16" / "Acme",
    )
    content["ats_score"] = "N/A"

    assert tracker.add_applied(content, force=False) is True

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    assert len(rows) == 1
    assert rows[0][4] == "N/A"


def test_add_applied_removes_manual_pending_row_first(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    folder = tmp_path / "Applications" / "2026-04-21" / "GammaInc"
    folder.mkdir(parents=True)
    url = "https://www.jobleads.com/pl/job/x--poland--aaa111deadbeef0000000000000000"
    assert tracker.add_manual_jobleads_pending(
        url=url, company="GammaInc", title="Dev", folder_abs=folder,
    ) is True

    content = _build_content(url, folder)
    assert tracker.add_applied(content, force=False) is True

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    assert len(rows) == 1
    assert rows[0][4] == "85%"
    assert rows[0][5] == url


def test_add_applied_converts_10_point_scale_to_percent(tmp_path, monkeypatch) -> None:
    tracker_path = tmp_path / "tracker.xlsx"
    monkeypatch.setattr(tracker, "TRACKER_PATH", tracker_path)

    content = _build_content(
        "https://example.com/jobs/5",
        tmp_path / "Applications" / "2026-04-16" / "Acme",
    )
    content["ats_score"] = "8/10"

    assert tracker.add_applied(content, force=False) is True

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    assert len(rows) == 1
    assert rows[0][4] == "80%"


def test_apply_pull_updates_updates_fields(tmp_path, monkeypatch) -> None:
    """apply_pull_updates writes Sent, Re-application, To Learn by ID."""
    from hunter import tracker
    from hunter.tracker import apply_pull_updates, add_applied

    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")

    # Write an applied row so we have something to update
    content = {
        "company_name": "PullCo",
        "job_title": "Frontend Dev",
        "stack": "Angular",
        "ats_score": "90",
        "apply_url": "https://example.com/pull/1",
        "output_folder": str(tmp_path / "PullCo"),
        "to_learn": "",
    }
    assert add_applied(content)

    # Find the row ID from tracker
    import openpyxl
    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    row_id = str(ws.cell(row=2, column=11).value or "").strip()
    wb.close()
    assert row_id

    # Now pull update with new Sent date and To Learn value
    updated_row = {
        "ID": row_id,
        "Sent": "2026-05-14",
        "Re-application": "+",
        "To Learn": "RxJS",
    }
    count = apply_pull_updates([updated_row])
    assert count == 1

    # Verify the change was written
    wb2 = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws2 = wb2.active
    row = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    wb2.close()
    assert row[7] == "2026-05-14"   # Sent (col 8, idx 7)
    assert row[8] == "+"            # Re-application (col 9, idx 8)
    assert row[9] == "RxJS"         # To Learn (col 10, idx 9)


def test_apply_pull_updates_noop_for_unknown_id(tmp_path, monkeypatch) -> None:
    from hunter import tracker
    from hunter.tracker import apply_pull_updates, add_applied

    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")

    content = {
        "company_name": "Ghost",
        "job_title": "Dev",
        "stack": "React",
        "ats_score": "70",
        "apply_url": "https://example.com/ghost/1",
        "output_folder": str(tmp_path / "Ghost"),
        "to_learn": "",
    }
    assert add_applied(content)

    count = apply_pull_updates([{"ID": "nonexistent", "Sent": "2026-05-14", "Re-application": "", "To Learn": ""}])
    assert count == 0
