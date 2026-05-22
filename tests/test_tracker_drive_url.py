"""Tests for tracker.get_drive_url_by_url and tracker.set_drive_url."""

import openpyxl
import pytest

from hunter import tracker
from hunter.tracker import COL_DRIVE_URL


def _make_tracker(tmp_path, rows: list[dict]) -> None:
    """Write a minimal tracker.xlsx with given rows (dicts with url, folder, drive_url)."""
    tracker_path = tmp_path / "tracker.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Company", "Job Title", "Stack", "ATS %", "URL",
               "Folder", "Sent", "Re-application", "To Learn", "ID", "Drive URL"])
    for r in rows:
        row = ["2026-05-22", r.get("company", "Acme"), r.get("title", "Dev"),
               "Angular", "85%", r.get("url", ""), r.get("folder", ""),
               "", "", "", r.get("id", "abc12345"), r.get("drive_url", "")]
        ws.append(row)
    wb.save(tracker_path)


# ---------------------------------------------------------------------------
# get_drive_url_by_url
# ---------------------------------------------------------------------------

def test_get_drive_url_returns_none_when_no_tracker(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1") is None


def test_get_drive_url_returns_none_when_url_not_found(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1"}])
    assert tracker.get_drive_url_by_url("https://example.com/jobs/99") is None


def test_get_drive_url_returns_none_when_col_empty(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1", "drive_url": ""}])
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1") is None


def test_get_drive_url_returns_stored_url(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    drive = "https://drive.google.com/drive/folders/abc"
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1", "drive_url": drive}])
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1") == drive


def test_get_drive_url_normalizes_job_url(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    drive = "https://drive.google.com/drive/folders/xyz"
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1", "drive_url": drive}])
    # URL with trailing slash and utm param — should still match
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1/?utm_source=test") == drive


# ---------------------------------------------------------------------------
# set_drive_url
# ---------------------------------------------------------------------------

def test_set_drive_url_noop_when_no_tracker(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    # Should not raise
    tracker.set_drive_url("https://example.com/jobs/1", "https://drive.google.com/x")


def test_set_drive_url_noop_when_url_not_found(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1"}])
    tracker.set_drive_url("https://example.com/jobs/99", "https://drive.google.com/x")
    # Original row should be untouched
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1") is None


def test_set_drive_url_writes_to_col_12(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1"}])
    drive = "https://drive.google.com/drive/folders/newid"
    tracker.set_drive_url("https://example.com/jobs/1", drive)

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    val = ws.cell(row=2, column=COL_DRIVE_URL).value
    wb.close()
    assert val == drive


def test_set_drive_url_is_idempotent(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    drive = "https://drive.google.com/drive/folders/abc"
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1", "drive_url": drive}])
    tracker.set_drive_url("https://example.com/jobs/1", drive)
    assert tracker.get_drive_url_by_url("https://example.com/jobs/1") == drive


def test_set_drive_url_updates_first_matching_row(tmp_path, monkeypatch):
    """When multiple rows share a URL (re-apply), only the first row is updated."""
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"url": "https://example.com/jobs/1", "id": "aaa11111"},
        {"url": "https://example.com/jobs/1", "id": "bbb22222"},
    ])
    drive = "https://drive.google.com/drive/folders/first"
    tracker.set_drive_url("https://example.com/jobs/1", drive)

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    row2_val = ws.cell(row=2, column=COL_DRIVE_URL).value
    row3_val = ws.cell(row=3, column=COL_DRIVE_URL).value
    wb.close()
    assert row2_val == drive
    assert not row3_val  # second row untouched


# ---------------------------------------------------------------------------
# read_all_tracker_rows includes Drive URL
# ---------------------------------------------------------------------------

def test_read_all_tracker_rows_includes_drive_url(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    drive = "https://drive.google.com/drive/folders/abc"
    _make_tracker(tmp_path, [{"url": "https://example.com/jobs/1", "drive_url": drive}])
    rows = tracker.read_all_tracker_rows()
    assert len(rows) == 1
    assert rows[0]["Drive URL"] == drive


def test_read_all_tracker_rows_drive_url_blank_for_old_rows(tmp_path, monkeypatch):
    """Rows without col 12 (old tracker format) return empty Drive URL."""
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    # Write an 11-column tracker (old format, no Drive URL col)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Company", "Job Title", "Stack", "ATS %", "URL",
               "Folder", "Sent", "Re-application", "To Learn", "ID"])
    ws.append(["2026-05-22", "Acme", "Dev", "Angular", "85%",
               "https://example.com/jobs/1", "Applications/x", "", "", "", "abc12345"])
    wb.save(tmp_path / "tracker.xlsx")

    rows = tracker.read_all_tracker_rows()
    assert rows[0]["Drive URL"] == ""
