"""Tests for /force two-step flow and cleanup logic."""

import asyncio
import pytest
from unittest.mock import AsyncMock, MagicMock, patch, call


def run(coro):
    """Helper: run a coroutine synchronously."""
    return asyncio.run(coro)


# ---------------------------------------------------------------------------
# tracker.delete_all_by_url
# ---------------------------------------------------------------------------

def _make_wb(rows):
    """Build a minimal openpyxl workbook for testing."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    # headers row
    ws.append(["Date", "Company", "Job Title", "Stack", "ATS %", "URL", "Folder",
               "Sent", "Re-app", "To Learn", "ID", "Drive URL"])
    for row in rows:
        ws.append(row)
    return wb


def test_delete_all_by_url_removes_all_statuses(tmp_path):
    """delete_all_by_url should delete FAIL, SKIP, MANUAL, and success rows."""
    import openpyxl
    tracker = tmp_path / "tracker.xlsx"

    wb = _make_wb([
        ["2026-05-10", "AcmeCorp", "Dev", "Angular", "95", "https://example.com/job/1",
         "Applications/2026-05-10/AcmeCorp", "", "", "", "aaa111", "https://drive.google.com/drive/folders/FOLDER1"],
        ["2026-05-11", "OtherCo",  "Dev", "React",   "87", "https://other.com/job/2",
         "Applications/2026-05-11/OtherCo", "", "", "", "bbb222", ""],
        ["2026-05-12", "AcmeCorp", "Dev", "Angular", "FAIL", "https://example.com/job/1",
         "", "", "", "", "ccc333", ""],
    ])
    wb.save(tracker)

    with patch("hunter.tracker.TRACKER_PATH", tracker):
        from hunter.tracker import delete_all_by_url
        result = delete_all_by_url("https://example.com/job/1")

    assert result["deleted"] == 2
    assert result["folder"] == "Applications/2026-05-10/AcmeCorp"
    assert result["drive_url"] == "https://drive.google.com/drive/folders/FOLDER1"

    # Only OtherCo row should remain
    wb2 = openpyxl.load_workbook(tracker)
    rows = list(wb2.active.iter_rows(min_row=2, values_only=True))
    companies = [r[1] for r in rows if any(r)]
    assert companies == ["OtherCo"]


def test_delete_all_by_url_unknown_url(tmp_path):
    """Returns zero deleted when URL not found."""
    tracker = tmp_path / "tracker.xlsx"
    wb = _make_wb([
        ["2026-05-10", "AcmeCorp", "Dev", "Angular", "95", "https://example.com/job/1",
         "Applications/AcmeCorp", "", "", "", "aaa111", ""],
    ])
    wb.save(tracker)

    with patch("hunter.tracker.TRACKER_PATH", tracker):
        from hunter.tracker import delete_all_by_url
        result = delete_all_by_url("https://other.com/job/999")

    assert result["deleted"] == 0
    assert result["folder"] is None
    assert result["drive_url"] is None


# ---------------------------------------------------------------------------
# tracker_cache.invalidate_url
# ---------------------------------------------------------------------------

def test_cache_invalidate_url_removes_from_all_indexes():
    from hunter.tracker_cache import TrackerCache

    async def _run():
        c = TrackerCache()
        row = {
            "ID": "abc123", "URL": "https://example.com/job/1",
            "Company": "AcmeCorp", "Job Title": "Dev",
            "ATS %": "95", "Sent": "", "Stack": "Angular",
            "Folder": "Applications/AcmeCorp", "Re-application": "",
            "To Learn": "", "Drive URL": "",
        }
        await c.add(row, sheet_row=2)
        assert await c.is_known_url("https://example.com/job/1")
        await c.invalidate_url("https://example.com/job/1")
        assert not await c.is_known_url("https://example.com/job/1")
        assert "abc123" not in c.rows
        assert "abc123" not in c.sheet_row_index

    run(_run())


def test_cache_invalidate_url_noop_unknown():
    from hunter.tracker_cache import TrackerCache

    async def _run():
        c = TrackerCache()
        await c.invalidate_url("https://notintracker.com/job/1")  # Should not raise

    run(_run())


# ---------------------------------------------------------------------------
# gdrive_client.folder_id_from_url
# ---------------------------------------------------------------------------

def test_folder_id_from_url_valid():
    from hunter.gdrive_client import folder_id_from_url
    url = "https://drive.google.com/drive/folders/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms"
    assert folder_id_from_url(url) == "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms"


def test_folder_id_from_url_invalid():
    from hunter.gdrive_client import folder_id_from_url
    assert folder_id_from_url("https://google.com") is None
    assert folder_id_from_url("") is None
    assert folder_id_from_url(None) is None


# ---------------------------------------------------------------------------
# _force_waiting state in telegram_bot
# ---------------------------------------------------------------------------

def _make_update(text: str, chat_id: int = 12345):
    update = MagicMock()
    update.message.text = text
    update.message.reply_text = AsyncMock()
    update.effective_chat.id = chat_id
    return update


def test_cmd_force_bare_adds_to_waiting():
    """Bare /force with no args → adds chat to _force_waiting, sends prompt."""
    import hunter.telegram_bot as bot

    async def _run():
        bot._force_waiting.clear()
        update = _make_update("/force")
        context = MagicMock()
        await bot.cmd_force(update, context)
        assert 12345 in bot._force_waiting
        update.message.reply_text.assert_called_once()
        msg = update.message.reply_text.call_args[0][0]
        assert "Force mode" in msg or "force" in msg.lower()

    run(_run())


def test_cmd_url_force_waiting_triggers_force_run():
    """Message from a chat in _force_waiting → calls _force_run, clears waiting."""
    import hunter.telegram_bot as bot

    async def _run():
        bot._force_waiting.clear()
        bot._force_waiting.add(12345)
        update = _make_update("https://justjoin.it/job-offer/some-company-dev")

        with patch.object(bot, "_force_run", new_callable=AsyncMock) as mock_run:
            context = MagicMock()
            await bot.cmd_url(update, context)

        assert 12345 not in bot._force_waiting
        mock_run.assert_called_once()

    run(_run())


def test_cmd_url_no_force_waiting_normal_flow():
    """Message from chat NOT in _force_waiting → normal flow (not _force_run)."""
    import hunter.telegram_bot as bot

    async def _run():
        bot._force_waiting.clear()
        update = _make_update("hello world")

        with patch.object(bot, "_force_run", new_callable=AsyncMock) as mock_run:
            with patch.object(bot, "_looks_like_paste", return_value=False):
                context = MagicMock()
                await bot.cmd_url(update, context)

        mock_run.assert_not_called()

    run(_run())


# ---------------------------------------------------------------------------
# _force_cleanup
# ---------------------------------------------------------------------------

def test_force_cleanup_deletes_server_folder(tmp_path):
    """_force_cleanup removes server folder when it exists."""
    import hunter.telegram_bot as bot

    async def _run():
        app_folder = tmp_path / "Applications" / "2026-05-22" / "AcmeCorp"
        app_folder.mkdir(parents=True)
        (app_folder / "CV.pdf").write_text("dummy")

        tracker_result = {"deleted": 1, "folder": str(app_folder), "drive_url": ""}
        update = _make_update("dummy")

        with patch("hunter.tracker.delete_all_by_url", return_value=tracker_result):
            with patch("hunter.tracker_cache.cache.invalidate_url", new_callable=AsyncMock):
                summary = await bot._force_cleanup("https://example.com/job/1", update)

        assert not app_folder.exists()
        assert "deleted" in summary.lower() or "AcmeCorp" in summary

    run(_run())


def test_force_cleanup_no_existing_entry():
    """_force_cleanup handles gracefully when URL is not in tracker."""
    import hunter.telegram_bot as bot

    async def _run():
        tracker_result = {"deleted": 0, "folder": None, "drive_url": None}
        update = _make_update("dummy")

        with patch("hunter.tracker.delete_all_by_url", return_value=tracker_result):
            with patch("hunter.tracker_cache.cache.invalidate_url", new_callable=AsyncMock):
                summary = await bot._force_cleanup("https://new-vacancy.com/job/99", update)

        assert "no existing" in summary.lower() or "0 row" in summary.lower() or "not found" in summary.lower()

    run(_run())
