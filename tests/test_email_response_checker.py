"""Tests for hunter/email_response_checker.py.

Test data based on real observed confirmation emails:
  - eRecruiter ATS (mail@stage.erecruiter.pl) used by NASK, EXATEL, Nexio, Medicover
  - SmartRecruiters (notification@smartrecruiters.com) used by Sigma Software
  - Direct company emails (inspeerity.com, consdata.com, etc.)
"""

import base64
import openpyxl
import pytest
from unittest.mock import MagicMock, patch

from hunter.email_response_checker import (
    ConfirmationEmail,
    MatchResult,
    _extract_body_text,
    _is_confirmation_subject,
    _message_date,
    _parse_erecruiter,
    _parse_smartrecruiters,
    _parse_direct,
    _parse_message,
    fetch_confirmation_emails,
    match_email,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _encode(text: str) -> str:
    return base64.urlsafe_b64encode(text.encode()).decode()


def _make_msg(
    subject: str,
    sender: str,
    body_text: str = "",
    internal_date_ms: int = 1716307200000,
) -> dict:
    parts = []
    if body_text:
        parts.append({
            "mimeType": "text/plain",
            "body": {"data": _encode(body_text)},
        })
    return {
        "internalDate": str(internal_date_ms),
        "payload": {
            "headers": [
                {"name": "Subject", "value": subject},
                {"name": "From", "value": sender},
            ],
            "parts": parts,
            "mimeType": "multipart/mixed" if parts else "text/plain",
            "body": {"data": _encode(body_text)} if not parts and body_text else {},
        },
    }


def _make_tracker(tmp_path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Date", "Company", "Job Title", "Stack", "ATS %", "URL",
        "Folder", "Sent", "Re-application", "To Learn", "ID", "Drive URL", "Response",
    ])
    for i, r in enumerate(rows):
        ws.append([
            "2026-05-22", r.get("company", "Acme"), r.get("title", "Dev"),
            "Angular", r.get("ats", "85%"), r.get("url", f"https://example.com/{i}"),
            "", "", "", "", r.get("id", f"id{i}"), "", r.get("response", ""),
        ])
    wb.save(tmp_path / "tracker.xlsx")


# ---------------------------------------------------------------------------
# _is_confirmation_subject — real-world subjects
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("subject", [
    # eRecruiter pattern (real)
    "NASK - Dziękujemy za złożenie aplikacji na stanowisko Senior Frontend Developer",
    "EXATEL - Dziękujemy za złożenie aplikacji na stanowisko Frontend Developer",
    # SmartRecruiters (real)
    "Thank you for applying to Sigma Software",
    # Generic English
    "Thanks for applying!",
    "Thank you for submitting your application",
    "Application submitted",
    "Application received",
    # Polish generic
    "Potwierdzenie aplikacji na stanowisko Angular Developer",
    "Dziękujemy za zainteresowanie ofertą",
])
def test_is_confirmation_subject_true(subject):
    assert _is_confirmation_subject(subject) is True


@pytest.mark.parametrize("subject", [
    "10 new Angular jobs for you",
    "New jobs matching your profile",
    "Nowe oferty pracy dla Ciebie",
    "Weekly jobs digest",
    "Your saved jobs this week",
])
def test_is_confirmation_subject_false(subject):
    assert _is_confirmation_subject(subject) is False


# ---------------------------------------------------------------------------
# _parse_erecruiter — real-world email format
# ---------------------------------------------------------------------------

def test_parse_erecruiter_nask():
    """Real NASK email via eRecruiter."""
    company, title = _parse_erecruiter(
        "NASK - Dziękujemy za złożenie aplikacji na stanowisko Senior Frontend Developer",
        "",
    )
    assert company == "NASK"
    assert title == "Senior Frontend Developer"


def test_parse_erecruiter_exatel():
    """Real EXATEL email via eRecruiter — title in body."""
    body = (
        "Dzień dobry,\n"
        "dziękujemy za złożenie aplikacji na stanowisko Frontend Developer i czas "
        "poświęcony na wypełnienie kwestionariuszy aplikacyjnych.\n"
    )
    company, title = _parse_erecruiter(
        "EXATEL - Dziękujemy za złożenie aplikacji na stanowisko Frontend Developer",
        body,
    )
    assert company == "EXATEL"
    assert title == "Frontend Developer"


def test_parse_erecruiter_body_fallback():
    """When subject doesn't match, extract title from body."""
    body = "dziękujemy za złożenie aplikacji na stanowisko Angular Developer i czas poświęcony."
    company, title = _parse_erecruiter("Dziękujemy za złożenie aplikacji", body)
    assert company == ""        # no company in subject without the dash pattern
    assert title == "Angular Developer"


def test_parse_erecruiter_no_match():
    company, title = _parse_erecruiter("Nowe oferty pracy", "")
    assert company == ""
    assert title == ""


# ---------------------------------------------------------------------------
# _parse_smartrecruiters — real-world email format
# ---------------------------------------------------------------------------

def test_parse_smartrecruiters_sigma_software():
    """Real Sigma Software email via SmartRecruiters."""
    body = (
        "Dear Ihar,\n\n"
        "Thank you for submitting your application for the position of "
        "Middle Front-End Developer. We strive to be an excellent workplace...\n"
    )
    company, title = _parse_smartrecruiters(
        "Thank you for applying to Sigma Software",
        body,
        "Sigma Software <notification@smartrecruiters.com>",
    )
    assert company == "Sigma Software"
    assert title == "Middle Front-End Developer"


def test_parse_smartrecruiters_company_from_subject():
    company, title = _parse_smartrecruiters(
        "Thank you for applying to Acme Corp",
        "",
        "Acme Corp <notification@smartrecruiters.com>",
    )
    assert company == "Acme Corp"


def test_parse_smartrecruiters_company_from_display_name():
    """When subject doesn't have company, use From display name."""
    company, title = _parse_smartrecruiters(
        "Thanks for applying!",
        "",
        "Acme Corp <notification@smartrecruiters.com>",
    )
    assert company == "Acme Corp"


def test_parse_smartrecruiters_no_title():
    company, title = _parse_smartrecruiters(
        "Thank you for applying to Sigma Software",
        "We received your application.",
        "Sigma Software <notification@smartrecruiters.com>",
    )
    assert company == "Sigma Software"
    assert title == ""


# ---------------------------------------------------------------------------
# _parse_direct — direct company emails
# ---------------------------------------------------------------------------

def test_parse_direct_polish_body():
    """Inspeerity-style email: Polish stanowisko in body."""
    body = "dziękujemy za złożenie aplikacji na stanowisko Angular Developer."
    company, title = _parse_direct(
        "Dziękujemy za aplikację",
        body,
        "Inspeerity <hr@inspeerity.com>",
    )
    assert title == "Angular Developer"
    assert company == "Inspeerity"


def test_parse_direct_company_from_display_name():
    company, title = _parse_direct(
        "Thank you for applying",
        "",
        "Creatio HR Team <hr@creatio.com>",
    )
    assert company == "Creatio HR Team"


def test_parse_direct_company_from_domain_fallback():
    """Generic sender name → extract company from domain."""
    company, title = _parse_direct(
        "Thank you for applying",
        "",
        "recruitment <hr@acme-company.com>",
    )
    # "recruitment" is in the skip list → domain fallback
    assert "acme" in company.lower()


def test_parse_direct_english_body_position():
    body = "Thank you for your interest. The position of Frontend Engineer is still open."
    company, title = _parse_direct(
        "Thank you for applying",
        body,
        "HR <hr@company.com>",
    )
    assert title == "Frontend Engineer"


# ---------------------------------------------------------------------------
# _extract_body_text
# ---------------------------------------------------------------------------

def test_extract_body_text_plain():
    payload = {
        "mimeType": "text/plain",
        "body": {"data": _encode("Hello world")},
    }
    assert _extract_body_text(payload) == "Hello world"


def test_extract_body_text_multipart():
    payload = {
        "parts": [
            {"mimeType": "text/plain", "body": {"data": _encode("Plain part")}},
            {"mimeType": "text/html", "body": {"data": _encode("<p>HTML</p>")}},
        ]
    }
    assert _extract_body_text(payload) == "Plain part"


def test_extract_body_text_empty():
    assert _extract_body_text({"mimeType": "text/html", "body": {}}) == ""


# ---------------------------------------------------------------------------
# _message_date
# ---------------------------------------------------------------------------

def test_message_date_returns_iso():
    msg = {"internalDate": "1716307200000"}
    assert _message_date(msg) == "2024-05-21"


# ---------------------------------------------------------------------------
# _parse_message — integration across parsers
# ---------------------------------------------------------------------------

def test_parse_message_erecruiter_nask():
    msg = _make_msg(
        subject="NASK - Dziękujemy za złożenie aplikacji na stanowisko Senior Frontend Developer",
        sender="eRecruiter <mail@stage.erecruiter.pl>",
    )
    result = _parse_message(msg)
    assert result is not None
    assert result.platform == "erecruiter"
    assert result.company == "NASK"
    assert result.title == "Senior Frontend Developer"


def test_parse_message_smartrecruiters():
    body = "Thank you for submitting your application for the position of Middle Front-End Developer."
    msg = _make_msg(
        subject="Thank you for applying to Sigma Software",
        sender="Sigma Software <notification@smartrecruiters.com>",
        body_text=body,
    )
    result = _parse_message(msg)
    assert result is not None
    assert result.platform == "smartrecruiters"
    assert result.company == "Sigma Software"
    assert result.title == "Middle Front-End Developer"


def test_parse_message_direct_company():
    body = "dziękujemy za złożenie aplikacji na stanowisko Angular Developer."
    msg = _make_msg(
        subject="Dziękujemy za złożenie aplikacji",
        sender="Inspeerity <hr@inspeerity.com>",
        body_text=body,
    )
    result = _parse_message(msg)
    assert result is not None
    assert result.platform == "direct"
    assert result.title == "Angular Developer"


def test_parse_message_skips_non_confirmation():
    msg = _make_msg(
        subject="10 new Angular jobs for you",
        sender="jobs-noreply@linkedin.com",
    )
    assert _parse_message(msg) is None


# ---------------------------------------------------------------------------
# match_email
# ---------------------------------------------------------------------------

def test_match_email_no_company(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    email = ConfirmationEmail(company="", title="Angular Dev", date="2026-05-20",
                              subject="...", platform="erecruiter")
    assert match_email(email).match_type == "no_match"


def test_match_email_exact(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "NASK", "title": "Senior Frontend Developer"}])
    email = ConfirmationEmail(company="NASK", title="Senior Frontend Developer",
                              date="2026-05-20", subject="...", platform="erecruiter")
    result = match_email(email)
    assert result.match_type == "exact"
    assert result.row_num == 2


def test_match_email_fuzzy_partial_title(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "NASK", "title": "Senior Frontend Developer"}])
    # "Frontend Developer" vs "Senior Frontend Developer" — "senior" is stop word → exact
    email = ConfirmationEmail(company="NASK", title="Frontend Developer",
                              date="2026-05-20", subject="...", platform="erecruiter")
    result = match_email(email)
    assert result.match_type in ("exact", "fuzzy")
    assert result.row_num == 2


def test_match_email_no_match_wrong_company(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Other Corp", "title": "Angular Developer"}])
    email = ConfirmationEmail(company="NASK", title="Angular Developer",
                              date="2026-05-20", subject="...", platform="erecruiter")
    assert match_email(email).match_type == "no_match"


def test_match_email_ambiguous(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"company": "Acme", "title": "Angular Developer", "id": "id0"},
        {"company": "Acme", "title": "Angular Engineer", "id": "id1"},
    ])
    email = ConfirmationEmail(company="Acme", title="Angular Dev",
                              date="2026-05-20", subject="...", platform="erecruiter")
    result = match_email(email)
    assert result.match_type == "ambiguous"
    assert result.row_num is None


def test_match_email_no_title_single_row(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Sigma Software", "title": "Middle Front-End Developer"}])
    # SmartRecruiters email where title couldn't be extracted
    email = ConfirmationEmail(company="Sigma Software", title="",
                              date="2026-05-20", subject="...", platform="smartrecruiters")
    result = match_email(email)
    assert result.match_type == "fuzzy"
    assert result.row_num == 2


def test_match_email_no_title_multiple_rows_ambiguous(tmp_path, monkeypatch):
    import hunter.tracker as t
    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"company": "Acme", "title": "Angular Developer", "id": "id0"},
        {"company": "Acme", "title": "React Developer", "id": "id1"},
    ])
    email = ConfirmationEmail(company="Acme", title="",
                              date="2026-05-20", subject="...", platform="direct")
    assert match_email(email).match_type == "ambiguous"


# ---------------------------------------------------------------------------
# fetch_confirmation_emails (mocked Gmail service)
# ---------------------------------------------------------------------------

def test_fetch_returns_empty_on_gmail_error():
    service = MagicMock()
    service.users().messages().list().execute.side_effect = Exception("network error")
    result = fetch_confirmation_emails(service, lookback_days=7)
    assert result == []


def test_fetch_filters_non_confirmation_subjects():
    msg = _make_msg(subject="10 new jobs for you", sender="jobs@erecruiter.pl")
    service = MagicMock()
    service.users().messages().list().execute.return_value = {"messages": [{"id": "1"}]}
    service.users().messages().get().execute.return_value = msg
    result = fetch_confirmation_emails(service, lookback_days=7)
    assert result == []


def test_fetch_parses_erecruiter_confirmation():
    msg = _make_msg(
        subject="NASK - Dziękujemy za złożenie aplikacji na stanowisko Senior Frontend Developer",
        sender="eRecruiter <mail@stage.erecruiter.pl>",
    )
    service = MagicMock()
    service.users().messages().list().execute.return_value = {"messages": [{"id": "1"}]}
    service.users().messages().get().execute.return_value = msg
    result = fetch_confirmation_emails(service, lookback_days=7)
    assert len(result) == 1
    assert result[0].company == "NASK"
    assert result[0].title == "Senior Frontend Developer"
    assert result[0].platform == "erecruiter"


def test_fetch_parses_smartrecruiters_confirmation():
    body = "Thank you for submitting your application for the position of Middle Front-End Developer."
    msg = _make_msg(
        subject="Thank you for applying to Sigma Software",
        sender="Sigma Software <notification@smartrecruiters.com>",
        body_text=body,
    )
    service = MagicMock()
    service.users().messages().list().execute.return_value = {"messages": [{"id": "1"}]}
    service.users().messages().get().execute.return_value = msg
    result = fetch_confirmation_emails(service, lookback_days=7)
    assert len(result) == 1
    assert result[0].company == "Sigma Software"
    assert result[0].title == "Middle Front-End Developer"


# ---------------------------------------------------------------------------
# run_confirmation_check — writes CONFIRMED to tracker
# ---------------------------------------------------------------------------

def test_run_writes_confirmed(tmp_path, monkeypatch):
    import hunter.tracker as t
    import hunter.email_response_checker as checker

    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "NASK", "title": "Senior Frontend Developer"}])

    confirmed_email = ConfirmationEmail(
        company="NASK", title="Senior Frontend Developer",
        date="2026-05-20", subject="NASK - Dziękujemy...", platform="erecruiter",
    )
    monkeypatch.setattr(checker, "fetch_confirmation_emails", lambda svc, days: [confirmed_email])

    with patch("hunter.email_response_checker.get_gmail_service", return_value=MagicMock()):
        results = checker.run_confirmation_check(lookback_days=7)

    assert results[0].match_type in ("exact", "fuzzy")
    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    from hunter.tracker import COL_RESPONSE
    assert ws.cell(row=2, column=COL_RESPONSE).value == "CONFIRMED"
    wb.close()


def test_run_does_not_overwrite_existing_response(tmp_path, monkeypatch):
    import hunter.tracker as t
    import hunter.email_response_checker as checker

    monkeypatch.setattr(t, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"company": "NASK", "title": "Senior Frontend Developer", "response": "INTERVIEW"}
    ])

    confirmed_email = ConfirmationEmail(
        company="NASK", title="Senior Frontend Developer",
        date="2026-05-20", subject="...", platform="erecruiter",
    )
    monkeypatch.setattr(checker, "fetch_confirmation_emails", lambda svc, days: [confirmed_email])

    with patch("hunter.email_response_checker.get_gmail_service", return_value=MagicMock()):
        checker.run_confirmation_check(lookback_days=7)

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    from hunter.tracker import COL_RESPONSE
    assert ws.cell(row=2, column=COL_RESPONSE).value == "INTERVIEW"  # untouched
    wb.close()
