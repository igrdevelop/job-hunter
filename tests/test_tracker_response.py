"""Tests for tracker Response column: set_response, lookup_by_company_and_title,
_title_tokens, _title_similarity."""

import openpyxl
import pytest

from hunter import tracker
from hunter.tracker import (
    COL_RESPONSE,
    _title_similarity,
    _title_tokens,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_tracker(tmp_path, rows: list[dict]) -> None:
    """Write a minimal tracker.xlsx. Supported row keys:
    company, title, ats, url, sent, response, id."""
    tracker_path = tmp_path / "tracker.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Date", "Company", "Job Title", "Stack", "ATS %", "URL",
        "Folder", "Sent", "Re-application", "To Learn", "ID", "Drive URL", "Response",
    ])
    for i, r in enumerate(rows):
        ws.append([
            "2026-05-22",
            r.get("company", "Acme"),
            r.get("title", "Developer"),
            "Angular",
            r.get("ats", "85%"),
            r.get("url", f"https://example.com/jobs/{i}"),
            "",
            r.get("sent", ""),
            "",
            "",
            r.get("id", f"abc1234{i}"),
            r.get("drive_url", ""),
            r.get("response", ""),
        ])
    wb.save(tracker_path)


# ---------------------------------------------------------------------------
# _title_tokens
# ---------------------------------------------------------------------------

def test_title_tokens_strips_stop_words():
    tokens = _title_tokens("Senior Angular Developer")
    assert "senior" not in tokens
    assert "angular" in tokens
    assert "developer" in tokens


def test_title_tokens_strips_diacritics():
    tokens = _title_tokens("Inżynier Frontend")
    assert "frontend" in tokens


def test_title_tokens_excludes_short_words():
    tokens = _title_tokens("UI JS Developer")
    # "ui" and "js" are len 2 — excluded
    assert "ui" not in tokens
    assert "js" not in tokens
    assert "developer" in tokens


def test_title_tokens_empty_string():
    assert _title_tokens("") == set()


# ---------------------------------------------------------------------------
# _title_similarity
# ---------------------------------------------------------------------------

def test_similarity_same_title():
    assert _title_similarity("Angular Developer", "Angular Developer") == 1.0


def test_similarity_senior_prefix_ignored():
    # "senior" is a stop word — tokens are the same
    score = _title_similarity("Senior Angular Developer", "Angular Developer")
    assert score == 1.0


def test_similarity_partial_overlap():
    # "Angular Developer" vs "Angular Engineer": {angular, developer} vs {angular, engineer}
    # intersection = {angular} / max(2, 2) = 0.5
    score = _title_similarity("Angular Developer", "Angular Engineer")
    assert score == pytest.approx(0.5)


def test_similarity_no_overlap():
    score = _title_similarity("Frontend Engineer", "Backend Java Developer")
    assert score == 0.0


def test_similarity_empty_titles():
    assert _title_similarity("", "Angular Developer") == 0.0
    assert _title_similarity("Angular Developer", "") == 0.0


# ---------------------------------------------------------------------------
# lookup_by_company_and_title
# ---------------------------------------------------------------------------

def test_lookup_returns_empty_when_no_tracker(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    assert tracker.lookup_by_company_and_title("Acme", "Angular Developer") == []


def test_lookup_returns_empty_when_company_not_found(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer"}])
    result = tracker.lookup_by_company_and_title("OtherCorp", "Angular Developer")
    assert result == []


def test_lookup_matches_same_company_same_title(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer"}])
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert len(result) == 1
    assert result[0]["company"] == "Acme"
    assert result[0]["title"] == "Angular Developer"
    assert result[0]["title_score"] == 1.0


def test_lookup_normalizes_company_legal_suffix(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme Sp. z o.o.", "title": "Angular Developer"}])
    # Email might have just "Acme" — normalize_company strips the suffix
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert len(result) == 1


def test_lookup_title_similarity_threshold(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"company": "Acme", "title": "Angular Developer"},
        {"company": "Acme", "title": "Java Backend Developer"},
    ])
    # "Angular Developer" should match the first row (score 1.0) but not the second (0.0)
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert len(result) == 1
    assert result[0]["title"] == "Angular Developer"


def test_lookup_respects_custom_threshold(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Engineer"}])
    # "Angular Developer" vs "Angular Engineer" = 0.5
    # With default threshold 0.5 → included
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer", title_min_score=0.5)
    assert len(result) == 1
    # With threshold 0.6 → excluded
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer", title_min_score=0.6)
    assert result == []


def test_lookup_sorted_by_score_descending(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [
        {"company": "Acme", "title": "Angular Engineer"},           # score 0.5
        {"company": "Acme", "title": "Senior Angular Developer"},   # score 1.0 (stop words stripped)
    ])
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer", title_min_score=0.4)
    assert result[0]["title"] == "Senior Angular Developer"
    assert result[1]["title"] == "Angular Engineer"


def test_lookup_returns_response_field(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer", "response": "CONFIRMED"}])
    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert result[0]["response"] == "CONFIRMED"


def test_lookup_response_empty_for_old_row(tmp_path, monkeypatch):
    """Rows without col 13 (old tracker format) return empty response."""
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Company", "Job Title", "Stack", "ATS %", "URL",
               "Folder", "Sent", "Re-application", "To Learn", "ID", "Drive URL"])
    ws.append(["2026-05-22", "Acme", "Angular Developer", "Angular", "85%",
               "https://example.com/jobs/1", "", "", "", "", "abc12345", ""])
    wb.save(tmp_path / "tracker.xlsx")

    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert len(result) == 1
    assert result[0]["response"] == ""


# ---------------------------------------------------------------------------
# set_response
# ---------------------------------------------------------------------------

def test_set_response_noop_when_no_tracker(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    tracker.set_response(2, "CONFIRMED")  # must not raise


def test_set_response_writes_confirmed(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer"}])
    tracker.set_response(2, "CONFIRMED")

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    val = ws.cell(row=2, column=COL_RESPONSE).value
    wb.close()
    assert val == "CONFIRMED"


def test_set_response_noop_for_out_of_range_row(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer"}])
    tracker.set_response(99, "CONFIRMED")  # row 99 doesn't exist — must not raise

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    assert ws.max_row == 2  # only header + 1 data row, nothing added


def test_set_response_overwrites_existing_value(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Dev", "response": "CONFIRMED"}])
    tracker.set_response(2, "INTERVIEW")

    wb = openpyxl.load_workbook(tmp_path / "tracker.xlsx", read_only=True, data_only=True)
    ws = wb.active
    val = ws.cell(row=2, column=COL_RESPONSE).value
    wb.close()
    assert val == "INTERVIEW"


def test_set_response_roundtrip_via_lookup(tmp_path, monkeypatch):
    monkeypatch.setattr(tracker, "TRACKER_PATH", tmp_path / "tracker.xlsx")
    _make_tracker(tmp_path, [{"company": "Acme", "title": "Angular Developer"}])
    tracker.set_response(2, "CONFIRMED")

    result = tracker.lookup_by_company_and_title("Acme", "Angular Developer")
    assert result[0]["response"] == "CONFIRMED"


# ---------------------------------------------------------------------------
# Schema: TRACKER_HEADERS includes Response
# ---------------------------------------------------------------------------

def test_tracker_headers_include_response():
    from hunter.tracker import TRACKER_HEADERS
    assert "Response" in TRACKER_HEADERS
    assert TRACKER_HEADERS.index("Response") == COL_RESPONSE - 1
