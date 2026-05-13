"""
tools/repair_tracker.py — Reconcile tracker.xlsx with Applications/ and to_send.xlsx.

What it does:
  1. Backs up tracker.xlsx before any changes
  2. Reads all existing tracker URLs (dedup key)
  3. Scans to_send.xlsx — adds any rows missing from tracker
  4. Scans Applications/{date}/{company}/content.json — adds missing rows
  5. Removes incomplete rows (no date AND no company AND no ATS)
  6. Sorts all data rows by Date ascending
  7. Saves tracker.xlsx

Run:
    python tools/repair_tracker.py          # dry-run — shows what would change
    python tools/repair_tracker.py --apply  # actually writes changes
"""

import json
import re
import shutil
import sys
import uuid

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
TRACKER_PATH = ROOT / "tracker.xlsx"
TO_SEND_PATH = ROOT / "to_send.xlsx"
APPLICATIONS_DIR = ROOT / "Applications"
BACKUP_DIR = ROOT / "backups"

DRY_RUN = "--apply" not in sys.argv

HEADERS = [
    "Date", "Company", "Job Title", "Stack",
    "ATS %", "URL", "Folder", "Sent", "Re-application", "To Learn", "ID",
]
# 1-based column indices
COL_DATE, COL_COMPANY, COL_TITLE, COL_STACK = 1, 2, 3, 4
COL_ATS, COL_URL, COL_FOLDER, COL_SENT = 5, 6, 7, 8
COL_REAPP, COL_TOLEARN, COL_ID = 9, 10, 11


def _new_id() -> str:
    return uuid.uuid4().hex[:8]


def _normalize_url(url: str) -> str:
    if not url:
        return ""
    url = url.strip().lower()
    url = re.sub(r"\?.*$", "", url)
    url = url.rstrip("/")
    return url


def _parse_ats(raw) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip().rstrip("%")
    try:
        return int(float(s))
    except ValueError:
        return None


def _relative_folder(folder_abs: str) -> str:
    """Convert absolute folder path to relative Applications/date/company form."""
    if not folder_abs:
        return ""
    p = Path(folder_abs)
    try:
        return str(p.relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        return folder_abs.replace("\\", "/")


def load_tracker_rows(ws) -> list[dict]:
    """Read all data rows from tracker worksheet into list of dicts."""
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rows.append({
            "row_num": i,
            "date": row[0],
            "company": row[1],
            "title": row[2],
            "stack": row[3],
            "ats": row[4],
            "url": row[5],
            "folder": row[6],
            "sent": row[7],
            "reapp": row[8],
            "tolearn": row[9],
            "id": row[10],
        })
    return rows


def load_to_send_rows() -> list[dict]:
    if not TO_SEND_PATH.exists():
        print(f"[repair] to_send.xlsx not found at {TO_SEND_PATH}")
        return []
    wb = openpyxl.load_workbook(TO_SEND_PATH)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({
            "date": row[0],
            "company": row[1],
            "title": row[2],
            "stack": row[3],
            "ats": row[4],
            "url": row[5],
            "folder": row[6],
            "sent": row[7],
            "reapp": row[8],
            "tolearn": row[9],
            "id": row[10] if len(row) > 10 else None,
        })
    print(f"[repair] to_send.xlsx: {len(rows)} rows")
    return rows


def scan_applications() -> list[dict]:
    """Scan Applications/{date}/{company}/content.json and build row dicts."""
    found = []
    if not APPLICATIONS_DIR.exists():
        return found

    for date_dir in sorted(APPLICATIONS_DIR.iterdir()):
        if not date_dir.is_dir():
            continue
        # Must look like a date 2026-xx-xx
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_dir.name):
            continue
        folder_date = date_dir.name

        for co_dir in sorted(date_dir.iterdir()):
            if not co_dir.is_dir():
                continue
            cj = co_dir / "content.json"
            if not cj.exists():
                continue
            try:
                data = json.loads(cj.read_text(encoding="utf-8"))
            except Exception as e:
                print(f"[repair] WARN: bad JSON in {cj}: {e}")
                continue

            url = (data.get("apply_url") or data.get("url") or "").strip()
            company = (data.get("company_name") or co_dir.name).strip()
            title = (data.get("job_title") or "").strip()
            stack = (data.get("stack") or "").strip()
            ats_raw = data.get("ats_score", "")
            folder_abs = str(co_dir)

            found.append({
                "date": folder_date,
                "company": company,
                "title": title,
                "stack": stack,
                "ats": str(ats_raw) if ats_raw else "",
                "url": url,
                "folder": folder_abs,
                "sent": None,
                "reapp": None,
                "tolearn": (data.get("to_learn") or ""),
                "id": None,  # will be assigned if missing
            })

    print(f"[repair] Applications/: {len(found)} content.json files")
    return found


def _row_is_garbage(r: dict) -> bool:
    """True if row has no meaningful data beyond a URL."""
    has_date = bool(r.get("date"))
    has_company = bool(r.get("company"))
    has_ats = bool(r.get("ats"))
    has_url = bool(r.get("url"))
    return has_url and not has_date and not has_company and not has_ats


def _sort_key(r: dict):
    d = r.get("date")
    if isinstance(d, datetime):
        return d.date().isoformat()
    if isinstance(d, date):
        return d.isoformat()
    return str(d or "0000-00-00")


def _apply_row_style(ws, row_num: int, values: list, ats_raw):
    row_font = Font(name="Calibri", size=11)

    # Determine fill
    ats_num = _parse_ats(ats_raw)
    status = str(ats_raw or "").upper()
    if status in ("SKIP", "REACT_SKIP"):
        fill = PatternFill("solid", fgColor="D9D9D9")
    elif status == "EXPIRED":
        fill = PatternFill("solid", fgColor="FCE4D6")
    elif status in ("FAIL", "MANUAL"):
        fill = PatternFill("solid", fgColor="FFC7CE")
    elif row_num % 2 == 0:
        fill = PatternFill("solid", fgColor="EEF2FA")
    else:
        fill = None

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = row_font
        if fill:
            cell.fill = fill

        if col == COL_ATS and ats_num is not None:
            cell.alignment = Alignment(horizontal="center")
            if ats_num >= 80:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(name="Calibri", size=11, color="276221", bold=True)
            elif ats_num >= 60:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
                cell.font = Font(name="Calibri", size=11, color="9C6500", bold=True)
            else:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(name="Calibri", size=11, color="9C0006", bold=True)

        if col == COL_URL and val:
            cell.hyperlink = str(val)
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

        if col == COL_SENT:
            cell.alignment = Alignment(horizontal="center")


def main():
    mode = "DRY RUN" if DRY_RUN else "APPLY"
    print(f"\n[repair] === tracker repair — {mode} ===\n")

    # ── Load current tracker ──────────────────────────────────────────────────
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active
    existing_rows = load_tracker_rows(ws)
    print(f"[repair] tracker.xlsx: {len(existing_rows)} data rows")

    # Build known URL set
    known_urls: set[str] = {
        _normalize_url(r["url"]) for r in existing_rows if r.get("url")
    }

    # Identify garbage rows (URL-only, no real data)
    garbage = [r for r in existing_rows if _row_is_garbage(r)]
    print(f"[repair] Garbage rows (URL-only): {len(garbage)}")

    # Keep only good rows
    good_rows = [r for r in existing_rows if not _row_is_garbage(r)]

    # ── Load to_send ──────────────────────────────────────────────────────────
    to_send_rows = load_to_send_rows()
    added_from_to_send = 0
    for r in to_send_rows:
        key = _normalize_url(r.get("url") or "")
        if key and key not in known_urls:
            good_rows.append(r)
            known_urls.add(key)
            added_from_to_send += 1
            print(f"  [+to_send] {r.get('date')} {r.get('company')} — {r.get('title')}")
    print(f"[repair] Added from to_send: {added_from_to_send}")

    # ── Scan Applications/ ────────────────────────────────────────────────────
    app_rows = scan_applications()
    added_from_apps = 0
    for r in app_rows:
        key = _normalize_url(r.get("url") or "")
        if not key:
            # No URL in content.json — use folder path as fallback key
            key = _normalize_url(r.get("folder") or "")
        if key and key not in known_urls:
            if not r.get("id"):
                r["id"] = _new_id()
            good_rows.append(r)
            known_urls.add(key)
            added_from_apps += 1
            print(f"  [+apps] {r.get('date')} {r.get('company')} — {r.get('title')}")
    print(f"[repair] Added from Applications/: {added_from_apps}")

    total_added = added_from_to_send + added_from_apps
    total_removed = len(garbage)

    print(f"\n[repair] Summary:")
    print(f"  Rows before:   {len(existing_rows)}")
    print(f"  Garbage removed: {total_removed}")
    print(f"  Added:           {total_added}")
    print(f"  Rows after:    {len(good_rows)}")

    if total_added == 0 and total_removed == 0:
        print("[repair] Nothing to do — tracker is already in sync.")
        return

    if DRY_RUN:
        print("\n[repair] Dry run — no changes written. Use --apply to save.")
        return

    # ── Backup ────────────────────────────────────────────────────────────────
    BACKUP_DIR.mkdir(exist_ok=True)
    backup_path = BACKUP_DIR / f"tracker_before_repair_{date.today().isoformat()}.xlsx"
    shutil.copy2(TRACKER_PATH, backup_path)
    print(f"\n[repair] Backup saved: {backup_path}")

    # ── Sort by date ──────────────────────────────────────────────────────────
    good_rows.sort(key=_sort_key)

    # ── Rebuild worksheet ─────────────────────────────────────────────────────
    # Clear all rows after header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.hyperlink = None
            cell.fill = PatternFill()
            cell.font = Font(name="Calibri", size=11)

    # Write sorted rows
    for i, r in enumerate(good_rows, start=2):
        folder_val = _relative_folder(str(r.get("folder") or ""))
        row_id = r.get("id") or _new_id()
        values = [
            r.get("date"),
            r.get("company"),
            r.get("title"),
            r.get("stack"),
            r.get("ats"),
            r.get("url"),
            folder_val,
            r.get("sent"),
            r.get("reapp"),
            r.get("tolearn"),
            row_id,
        ]
        _apply_row_style(ws, i, values, r.get("ats"))

    # ── Column widths (keep existing or set sensible defaults) ────────────────
    col_widths = {1: 12, 2: 22, 3: 36, 4: 20, 5: 8, 6: 55, 7: 45, 8: 8, 9: 12, 10: 35, 11: 12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(TRACKER_PATH)
    print(f"[repair] tracker.xlsx saved — {len(good_rows)} rows, sorted by date.")


if __name__ == "__main__":
    main()
