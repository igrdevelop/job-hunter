"""
One-off / maintenance: remove duplicate rows in tracker.xlsx by normalized URL.
Keeps the best row per URL: highest ATS quality, then Sent filled, then earliest row.

Usage: python tools/dedupe_tracker.py
"""
from __future__ import annotations

import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

# Repo root
ROOT = Path(__file__).resolve().parent.parent
import sys

sys.path.insert(0, str(ROOT))

from hunter.config import TRACKER_PATH
from hunter.tracker import normalize_url

URL_COL = 6
ATS_COL = 5
SENT_COL = 8


def _row_score(ws, row_num: int) -> tuple:
    ats = str(ws.cell(row=row_num, column=ATS_COL).value or "").strip()
    sent = str(ws.cell(row=row_num, column=SENT_COL).value or "").strip()
    # Higher = better candidate to keep
    tier = 0
    if ats.upper() in ("FAIL", "SKIP"):
        tier = 0
    elif ats in ("", "?"):
        tier = 1
    else:
        tier = 2  # looks like a real ATS % or status
    sent_bonus = 1 if sent else 0
    # Prefer lower row number on tie (earlier in file)
    return (tier, sent_bonus, -row_num)


def main() -> None:
    if not TRACKER_PATH.exists():
        print(f"No tracker at {TRACKER_PATH}")
        return

    backup = TRACKER_PATH.with_suffix(
        f".backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(TRACKER_PATH, backup)
    print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active

    groups: dict[str, list[int]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=r, column=URL_COL).value
        if not url_cell:
            continue
        nu = normalize_url(str(url_cell).strip())
        if not nu:
            continue
        groups[nu].append(r)

    to_delete: list[int] = []
    for nu, rows in groups.items():
        if len(rows) <= 1:
            continue
        best = max(rows, key=lambda rn: _row_score(ws, rn))
        for rn in rows:
            if rn != best:
                to_delete.append(rn)

    to_delete.sort(reverse=True)
    for rn in to_delete:
        ws.delete_rows(rn)

    wb.save(TRACKER_PATH)
    wb.close()
    print(f"Removed {len(to_delete)} duplicate row(s). Kept one row per URL.")


if __name__ == "__main__":
    main()
