"""
tools/fix_pracuj_urls.py — Fix broken pracuj.pl URLs in tracker.xlsx.

Finds rows where URL is https://www.pracuj.pl/praca/oferta,ID (no slug),
reconstructs the URL using the Job Title column, and updates the cell.

Usage:
    python tools/fix_pracuj_urls.py           # dry-run — shows what would change
    python tools/fix_pracuj_urls.py --apply   # actually update file
"""

import re
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font

TRACKER_PATH  = ROOT / "tracker.xlsx"
APPLY = "--apply" in sys.argv

BROKEN_URL_RE = re.compile(r"^https://www\.pracuj\.pl/praca/oferta,(\d+)$")


def slugify(title: str) -> str:
    slug = title.lower().strip()
    slug = re.sub(r"[^\w\s-]", "", slug)
    slug = re.sub(r"[\s_]+", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug)
    return slug.strip("-")


def fix_file(path: Path) -> int:
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    try:
        url_col   = headers.index("URL") + 1
        title_col = headers.index("Job Title") + 1
    except ValueError as e:
        print(f"  [skip] {path.name}: column not found: {e}")
        return 0

    fixed = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        url   = str(ws.cell(row, url_col).value or "").strip()
        title = str(ws.cell(row, title_col).value or "").strip()

        m = BROKEN_URL_RE.match(url)
        if not m:
            continue

        offer_id = m.group(1)

        if not title:
            print(f"    [{row:3}] ⚠️  No title — cannot fix: {url}")
            skipped += 1
            continue

        slug = slugify(title)
        new_url = f"https://www.pracuj.pl/praca/{slug},oferta,{offer_id}"

        print(f"    [{row:3}] {title[:50]}")
        print(f"           OLD: {url}")
        print(f"           NEW: {new_url}")

        if APPLY:
            cell = ws.cell(row, url_col, value=new_url)
            cell.hyperlink = new_url
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

        fixed += 1

    print(f"  → Fixable: {fixed}  |  No title: {skipped}")

    if APPLY and fixed:
        wb.save(path)
        print(f"  ✅ {path.name} saved.")

    return fixed


def main():
    mode = "APPLY" if APPLY else "DRY RUN"
    print(f"\n[fix_pracuj] === Fix broken pracuj.pl URLs — {mode} ===\n")

    print(f"📋 tracker.xlsx:")
    total = fix_file(TRACKER_PATH)
    print(f"\n{'='*60}")
    if total == 0:
        print("[fix_pracuj] Nothing to fix.")
    elif not APPLY:
        print(f"[fix_pracuj] {total} URLs to fix. Run with --apply to save.")
    else:
        print(f"[fix_pracuj] ✅ Done — {total} URLs fixed.")


if __name__ == "__main__":
    main()

