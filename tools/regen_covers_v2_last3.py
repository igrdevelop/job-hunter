#!/usr/bin/env python3
"""Regenerate cover letters (EN+PL) for the last N tracker rows with folders.

Writes Cover_Letter_EN_v2.docx/pdf and Cover_Letter_PL_v2.docx/pdf next to originals.
Does not modify content.json or existing Cover_Letter_* without _v2.

Usage (from repo root):
  python tools/regen_covers_v2_last3.py
  python tools/regen_covers_v2_last3.py --count 3
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

# Repo root (parent of tools/)
PROJECT_DIR = Path(__file__).resolve().parent.parent
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from hunter.config import LLM_API_KEY, LLM_MODEL, LLM_PROVIDER, TRACKER_PATH  # noqa: E402
from llm_client import LLMError, call_llm  # noqa: E402


def _resolve_folder(raw: str) -> Path:
    p = Path(raw.strip())
    if p.is_absolute():
        return p
    return PROJECT_DIR / p


def _soffice() -> Path:
    return Path(r"C:\Program Files\LibreOffice\program\soffice.exe")


def _convert_docx_pdf(docx: Path) -> None:
    exe = _soffice()
    if not exe.is_file():
        print(f"  [WARN] LibreOffice not found at {exe}, skip PDF for {docx.name}")
        return
    r = subprocess.run(
        [str(exe), "--headless", "--convert-to", "pdf", "--outdir", str(docx.parent), str(docx)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if r.returncode != 0:
        print(f"  [WARN] PDF convert failed: {r.stderr.strip()[:200]}")


def _load_system_prompt() -> str:
    prompt_path = PROJECT_DIR / "prompts" / "system_prompt.md"
    profile_path = PROJECT_DIR / "prompts" / "candidate_profile.md"
    instructions = prompt_path.read_text(encoding="utf-8")
    if profile_path.exists():
        return profile_path.read_text(encoding="utf-8") + "\n\n---\n\n" + instructions
    return instructions


def _last_tracker_folders(count: int) -> list[tuple[int, str, str]]:
    """Return up to `count` items: (excel_row, folder_path_str, url) from bottom of sheet."""
    if not TRACKER_PATH.is_file():
        raise SystemExit(f"tracker not found: {TRACKER_PATH}")
    wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
    ws = wb.active
    max_row = ws.max_row or 2
    out: list[tuple[int, str, str]] = []
    seen: set[str] = set()
    for r in range(max_row, 1, -1):
        row = [ws.cell(r, c).value for c in range(1, 12)]
        if not row or len(row) < 7:
            continue
        folder_raw = str(row[6] or "").strip()
        url = str(row[5] or "").strip()
        if not folder_raw:
            continue
        folder_abs = _resolve_folder(folder_raw)
        if not (folder_abs / "content.json").is_file():
            continue
        key = str(folder_abs.resolve())
        if key in seen:
            continue
        seen.add(key)
        out.append((r, folder_raw, url))
        if len(out) >= count:
            break
    wb.close()
    return out


def _regen_covers_for_folder(folder_abs: Path, system_prompt: str, url: str) -> None:
    content_path = folder_abs / "content.json"
    job_path = folder_abs / "job_posting.txt"
    content = json.loads(content_path.read_text(encoding="utf-8"))
    job_text = job_path.read_text(encoding="utf-8") if job_path.is_file() else ""

    ctx = {
        "company_name": content.get("company_name"),
        "job_title": content.get("job_title"),
        "stack": content.get("stack"),
        "lang": content.get("lang"),
        "resume_summary": (content.get("resume_en") or {}).get("summary"),
        "prior_cover_letter_en": content.get("cover_letter_en"),
        "prior_cover_letter_pl": content.get("cover_letter_pl"),
    }

    user_message = (
        "Regenerate ONLY the cover letters for this application.\n\n"
        "Return a JSON object with EXACTLY these two string keys:\n"
        '  "cover_letter_en"\n'
        '  "cover_letter_pl"\n\n'
        "Follow ALL **Cover Letter EN** and **Cover Letter PL** rules in your instructions "
        "(Dear Hiring Manager, blank line, 3-4 body paragraphs separated by \\n\\n, no signature block, "
        "metrics, posting anchor, hybrid CTA).\n\n"
        f"Job posting:\n{job_text or '(missing job_posting.txt — use URL and context only)'}\n\n"
        f"Original URL: {url or '(unknown)'}\n\n"
        "Use this context for facts and tone (do not copy old letters verbatim; improve structure):\n"
        f"{json.dumps(ctx, ensure_ascii=False, indent=2)}"
    )

    if not LLM_API_KEY:
        raise SystemExit("LLM_API_KEY / ANTHROPIC_API_KEY not set")

    print(f"  LLM {LLM_PROVIDER}/{LLM_MODEL} …")
    try:
        result = call_llm(
            system_prompt=system_prompt,
            user_message=user_message,
            provider=LLM_PROVIDER,
            model=LLM_MODEL,
            api_key=LLM_API_KEY,
            max_tokens=8192,
        )
    except LLMError as e:
        print(f"  [ERROR] LLM: {e}")
        return

    en = result.get("cover_letter_en")
    pl = result.get("cover_letter_pl")
    if not isinstance(en, str) or len(en.strip()) < 80:
        print("  [ERROR] Invalid cover_letter_en from model")
        return
    if not isinstance(pl, str) or len(pl.strip()) < 80:
        print("  [WARN] Weak cover_letter_pl, keeping EN-only for PL file check")
        pl = pl if isinstance(pl, str) else ""

    work = dict(content)
    work["cover_letter_en"] = en.strip()
    if pl:
        work["cover_letter_pl"] = pl.strip()

    from apply_agent import _cover_letter_review_loop  # noqa: E402

    work = _cover_letter_review_loop(work)

    from docx import Document  # noqa: E402
    from generate_docs import build_cover_letter, set_margins, set_author  # noqa: E402

    def _save_cl_docx(text: str, name: str) -> Path:
        doc = Document()
        set_margins(doc)
        build_cover_letter(doc, text)
        out = folder_abs / name
        set_author(doc)
        doc.save(out)
        print(f"  [OK] {out.name}")
        return out

    en_docx = _save_cl_docx(work["cover_letter_en"], "Cover_Letter_EN_v2.docx")
    if work.get("cover_letter_pl"):
        pl_docx = _save_cl_docx(work["cover_letter_pl"], "Cover_Letter_PL_v2.docx")
        _convert_docx_pdf(pl_docx)
    _convert_docx_pdf(en_docx)

    snap = {
        "cover_letter_en": work["cover_letter_en"],
        "cover_letter_pl": work.get("cover_letter_pl", ""),
        "source_url": url,
    }
    (folder_abs / "cover_letters_v2_snapshot.json").write_text(
        json.dumps(snap, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  [OK] cover_letters_v2_snapshot.json")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--count", type=int, default=3, help="How many latest tracker rows")
    args = parser.parse_args()

    system_prompt = _load_system_prompt()
    rows = _last_tracker_folders(args.count)
    if not rows:
        print("No suitable tracker rows (need Folder + content.json).")
        return

    print(f"Regenerating covers for {len(rows)} folder(s):\n")
    for excel_row, folder_raw, url in rows:
        folder_abs = _resolve_folder(folder_raw)
        print(f"Row {excel_row}: {folder_abs}")
        _regen_covers_for_folder(folder_abs, system_prompt, url)
        print()


if __name__ == "__main__":
    main()
